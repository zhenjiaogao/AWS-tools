#!/usr/bin/env python3
"""
支持获取指定区域的RDS MySQL和RDS PostgreSQL实例信息，并计算RDS和Aurora的RI定价
Aurora转换支持r7g、r8g两种实例类型，分别计算实例成本、存储成本和总MRR成本
RDS替换支持m7g、m8g、r7g、r8g四种实例类型，分别计算实例成本、存储成本和总MRR成本

Aurora Provisioned Graviton换代评估：
- 扫描现有 Aurora MySQL / Aurora PostgreSQL provisioned 实例
- 评估从当前代次升级到 r7g、r8g 的 instance cost 节省（只算实例成本，不含存储）
- 支持 r5/r6g/r6i/r7g 等代次向上升级映射

支持的引擎类型：
1. RDS MySQL -> Aurora MySQL
2. RDS PostgreSQL -> Aurora PostgreSQL
3. Aurora MySQL Provisioned Graviton换代评估
4. Aurora PostgreSQL Provisioned Graviton换代评估

支持的实例类型：
1. M系列实例：迁移Graviton3/4时使用m7g、m8g，迁移Aurora时使用r7g、r8g
2. R系列实例：迁移Graviton3/4时使用r7g、r8g，迁移Aurora时使用r7g、r8g
3. C系列实例：迁移Graviton3/4时使用m7g、m8g，迁移Aurora时使用r7g、r8g

Multi-AZ存储成本优化：
1. 当RDS是Multi-AZ架构时，迁移到Aurora时，存储成本不乘以2，仅算一份
2. 当Multi-AZ架构并且有多个read replica时，迁移到Aurora的存储成本计算到Primary节点上，仅计算一份
3. RDS替换场景下，Multi-AZ存储成本仍然乘以2（保持RDS原有逻辑）

RDS集群支持：
1. 通过source_db_instance_identifier识别RDS集群关系
2. 当read replica的source_db_instance_identifier值和primary的db_instance_identifier值相同时，识别为集群
3. 迁移到Aurora时，集群的存储成本仅计算一次，并加和到Primary节点上
4. Read replica节点的Aurora存储成本为0，避免重复计算
5. 支持复杂的集群拓扑，包括一个primary对应多个read replica的场景

Aurora存储容量优化：
1. aurora_allocate_storage_gb 基于 Primary节点的 used_storage_gb（实际使用量）计算
2. 符合Aurora共享存储特性，不累加集群中所有节点的存储量
3. 更准确反映迁移到Aurora后的实际存储需求和成本，通过 API 获取 standard 价格
4. 优先使用Primary节点的实际使用量，如果未获取到则回退到分配量
5. 避免基于过度分配或重复计算的存储容量进行成本计算

提升执行效率：
1. 在脚本中实现了并发获取 cloudwatch 信息等，使用 --optimize 启用并发处理，-b -w 指定 batch size 以及并发数
2. 将 RDS 和 Aurora 的 Pricing 的获取，统一使用一个方法

增加区域兼容性验证：该脚本仅支持在 global region 执行

使用方式：
  python rds_aurora_multi_generation_pricing_analyzer.py <region>                    # 分析MySQL和PostgreSQL
  python rds_aurora_multi_generation_pricing_analyzer.py <region> --engine mysql      # 仅分析MySQL
  python rds_aurora_multi_generation_pricing_analyzer.py <region> --engine postgresql # 仅分析PostgreSQL
"""

import argparse
import csv
import json
import logging
import sys
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import boto3
from botocore.exceptions import ClientError, NoCredentialsError
import pandas as pd


class CompleteUnifiedRDSAuroraAnalyzer:
    def __init__(self, region: str, enable_optimization: bool = False, max_workers: int = 10, batch_size: int = 50, engine: str = 'all', mode: str = 'all'):
        """
        完整统一分析器，包含原始脚本的所有功能
        engine: 'mysql', 'postgresql', 'all'
        mode: 'rds-to-aurora', 'rds-replacement', 'aurora-replacement', 'all'
        """
        self.region = region
        self.enable_optimization = enable_optimization
        self.max_workers = max_workers if enable_optimization else 1
        self.batch_size = batch_size
        self.engine = engine  # 'mysql', 'postgresql', 'all'
        self.mode = mode  # 'rds-to-aurora', 'rds-replacement', 'aurora-replacement', 'all'
        
        # 设置基础日志（用于兼容性检查）
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
        # 区域兼容性检查
        if not self._check_region_compatibility():
            print("\n❌ 脚本因区域兼容性问题终止执行")
            sys.exit(1)
        
        # AWS客户端
        self.rds_client = boto3.client('rds', region_name=region)
        self.pricing_client = boto3.client('pricing', region_name='us-east-1')
        self.sts_client = boto3.client('sts', region_name=region)
        self.cloudwatch_client = boto3.client('cloudwatch', region_name=region)
        
        # MySQL 缓存
        self.pricing_cache = {}
        self.aurora_pricing_cache = {}
        self.storage_pricing_cache = {}
        self.aurora_storage_pricing_cache = {}
        
        # PostgreSQL 缓存
        self.pg_pricing_cache = {}
        self.aurora_pg_pricing_cache = {}
        self.pg_storage_pricing_cache = {}
        self.aurora_pg_storage_pricing_cache = {}
        
        # 实例类型配置
        self.aurora_generations = ['r6g', 'r7g', 'r8g']
        self.rds_replacement_generations = ['r6g', 'm6g', 'r7g', 'm7g', 'r8g', 'm8g']
        
        # Aurora Graviton换代目标代次（只算instance cost, IO-Optimized模式 = Standard RI × 1.3）
        self.aurora_graviton_upgrade_targets = ['r6g', 'r7g', 'r8g']
        
        # Aurora Provisioned 实例分析结果缓存
        self.aurora_provisioned_results = []
        
        self._setup_logging()

    def _setup_logging(self):
        """设置日志"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'{self.region}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def _print_progress(self, message: str):
        """打印进度信息"""
        pass

    def get_account_id(self) -> str:
        """获取AWS账户ID"""
        try:
            response = self.sts_client.get_caller_identity()
            return response['Account']
        except Exception as e:
            self.logger.error(f"获取账户ID失败: {e}")
            return "unknown"

    def _get_pricing_location(self) -> str:
        """获取定价API的区域名称"""
        region_mapping = {
            "us-east-1": "US East (N. Virginia)",
            "us-east-2": "US East (Ohio)",
            "us-west-1": "US West (N. California)",
            "us-west-2": "US West (Oregon)",
            "ap-east-1": "Asia Pacific (Hong Kong)",
            "ap-south-1": "Asia Pacific (Mumbai)",
            "ap-south-2": "Asia Pacific (Hyderabad)",
            "ap-northeast-1": "Asia Pacific (Tokyo)",
            "ap-northeast-2": "Asia Pacific (Seoul)",
            "ap-northeast-3": "Asia Pacific (Osaka)",
            "ap-southeast-1": "Asia Pacific (Singapore)",
            "ap-southeast-2": "Asia Pacific (Sydney)",
            "ap-southeast-3": "Asia Pacific (Jakarta)",
            "ap-southeast-4": "Asia Pacific (Melbourne)",
            "ca-central-1": "Canada (Central)",
            "eu-central-1": "EU (Frankfurt)",
            "eu-central-2": "EU (Zurich)",
            "eu-west-1": "EU (Ireland)",
            "eu-west-2": "EU (London)",
            "eu-west-3": "EU (Paris)",
            "eu-north-1": "EU (Stockholm)",
            "eu-south-1": "EU (Milan)",
            "eu-south-2": "EU (Spain)",
            "me-south-1": "Middle East (Bahrain)",
            "me-central-1": "Middle East (UAE)",
            "af-south-1": "Africa (Cape Town)",
            "sa-east-1": "South America (Sao Paulo)"
        }
        return region_mapping.get(self.region, self.region)

    def _get_current_ec2_region(self):
        """获取当前EC2实例所在的区域"""
        try:
            # 尝试IMDSv2
            token_response = requests.put(
                'http://169.254.169.254/latest/api/token',
                headers={'X-aws-ec2-metadata-token-ttl-seconds': '21600'},
                timeout=2
            )
            if token_response.status_code == 200:
                token = token_response.text
                region_response = requests.get(
                    'http://169.254.169.254/latest/meta-data/placement/region',
                    headers={'X-aws-ec2-metadata-token': token},
                    timeout=2
                )
                if region_response.status_code == 200:
                    return region_response.text.strip()
        except:
            pass
        
        try:
            # 尝试IMDSv1
            response = requests.get(
                'http://169.254.169.254/latest/meta-data/placement/region',
                timeout=2
            )
            if response.status_code == 200:
                return response.text.strip()
        except:
            pass
        
        try:
            session = boto3.Session()
            return session.region_name
        except:
            pass
            
        # 尝试通过STS推断区域
        try:
            sts_client = boto3.client('sts')
            identity = sts_client.get_caller_identity()
            arn = identity.get('Arn', '')
            if ':cn-' in arn:
                return 'cn-north-1'  # 默认中国区域
            else:
                return 'us-east-1'  # 默认全球区域
        except:
            return None

    def _is_china_region(self, region):
        """判断是否为中国区域"""
        return region.startswith('cn-') if region else False

    def _check_region_compatibility(self):
        """检查区域兼容性"""
        current_region = self._get_current_ec2_region()
        
        if not current_region:
            print("⚠️  警告: 无法确定当前EC2实例所在区域，跳过兼容性检查")
            return True
        
        current_is_china = self._is_china_region(current_region)
        target_is_china = self._is_china_region(self.region)
        
        # 当前EC2在中国区域时，直接提示不支持
        if current_is_china:
            print(f"""
❌ 不支持的运行环境！

当前EC2实例位于: {current_region} (中国区域)
此脚本不支持在中国区域运行。
""")
            return False
        
        # 当前EC2在全球区域，但目标是中国区域时的提示
        elif not current_is_china and target_is_china:
            print(f"""
❌ 区域兼容性错误！

当前EC2实例位于: {current_region} (全球区域)
目标评估区域: {self.region} (中国区域)

⚠️  在全球区域的EC2实例上无法评估中国区域的RDS实例！
请在中国区域的EC2实例上运行此脚本来分析中国区域的RDS实例。
""")
            return False
        
        # 全球区域到全球区域
        print(f"✅ 区域兼容性检查通过: {current_region} -> {self.region} (全球分区)")
        return True

    def _should_process_instance(self, instance_class: str) -> bool:
        """判断是否应该处理该实例类型"""
        if not instance_class.startswith('db.'):
            return False
        
        instance_family = instance_class.split('.')[1] if len(instance_class.split('.')) > 1 else ''
        
        if instance_family.startswith(('m', 'r', 'c')):
            return True
        else:
            self.logger.warning(f"未知实例族类型: {instance_class}，将继续处理")
            return True

    def get_storage_info(self, db_instance_identifier: str, db_instance_data: Dict) -> Dict:
        """获取存储信息"""
        storage_info = {
            'allocated_storage_gb': db_instance_data.get('AllocatedStorage', 0),
            'storage_type': db_instance_data.get('StorageType', 'gp2'),
            'iops': db_instance_data.get('Iops', 0),
            'storage_throughput': db_instance_data.get('StorageThroughput', 0),
            'used_storage_gb': 0,
            'free_storage_gb': 0,
            'storage_utilization_percent': 0
        }
        
        # 获取CloudWatch指标
        try:
            used_storage = self._get_used_storage_from_cloudwatch(db_instance_identifier)
            if used_storage is not None:
                storage_info['used_storage_gb'] = used_storage
                storage_info['free_storage_gb'] = max(0, storage_info['allocated_storage_gb'] - used_storage)
                if storage_info['allocated_storage_gb'] > 0:
                    storage_info['storage_utilization_percent'] = (used_storage / storage_info['allocated_storage_gb']) * 100
        except Exception as e:
            self.logger.warning(f"获取CloudWatch指标失败: {e}")
        
        self.logger.info(f"实例 {db_instance_identifier} 存储信息: 类型={storage_info['storage_type']}, "
                        f"分配={storage_info['allocated_storage_gb']}GB, "
                        f"剩余={storage_info['free_storage_gb']}GB, "
                        f"IOPS={storage_info['iops']}, "
                        f"吞吐量={storage_info['storage_throughput']}MB/s")
        
        return storage_info

    def _get_used_storage_from_cloudwatch(self, db_instance_identifier: str) -> Optional[float]:
        """从CloudWatch获取已使用存储空间"""
        try:
            end_time = time.time()
            start_time = end_time - 86400  # 24小时前
            
            response = self.cloudwatch_client.get_metric_statistics(
                Namespace='AWS/RDS',
                MetricName='FreeStorageSpace',
                Dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_identifier}],
                StartTime=start_time,
                EndTime=end_time,
                Period=3600,
                Statistics=['Average']
            )
            
            if response['Datapoints']:
                # 获取最新的空闲存储空间（字节）
                latest_free_bytes = min(dp['Average'] for dp in response['Datapoints'])
                
                # 获取分配的存储空间
                db_response = self.rds_client.describe_db_instances(
                    DBInstanceIdentifier=db_instance_identifier
                )
                allocated_gb = db_response['DBInstances'][0]['AllocatedStorage']
                allocated_bytes = allocated_gb * 1024 * 1024 * 1024
                
                # 计算已使用存储空间
                used_bytes = allocated_bytes - latest_free_bytes
                used_gb = used_bytes / (1024 * 1024 * 1024)
                
                return max(0, used_gb)
            
            return None
            
        except Exception as e:
            self.logger.warning(f"获取CloudWatch存储指标失败: {e}")
            return None

    def get_cloudwatch_metrics(self, db_instance_identifier: str) -> Dict:
        """获取CloudWatch指标"""
        metrics = {
            'iops-avg.1Hr/15-day period': 0,
            'cpu-max.1Hr/15-day period': 0,
            'cpu-min.1Hr/15-day period': 0,
            'cpu-avg.1Hr/15-day period': 0
        }
        
        try:
            # 15天前到现在的时间范围
            end_time = time.time()
            start_time = end_time - (15 * 24 * 3600)  # 15天前
            
            # IOPS指标（ReadIOPS + WriteIOPS）
            try:
                read_iops_response = self.cloudwatch_client.get_metric_statistics(
                    Namespace='AWS/RDS',
                    MetricName='ReadIOPS',
                    Dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_identifier}],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=3600,
                    Statistics=['Average']
                )
                
                write_iops_response = self.cloudwatch_client.get_metric_statistics(
                    Namespace='AWS/RDS',
                    MetricName='WriteIOPS',
                    Dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_identifier}],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=3600,
                    Statistics=['Average']
                )
                
                # 计算总IOPS平均值
                if read_iops_response['Datapoints'] and write_iops_response['Datapoints']:
                    # 按时间戳对齐数据点
                    read_dict = {dp['Timestamp']: dp['Average'] for dp in read_iops_response['Datapoints']}
                    write_dict = {dp['Timestamp']: dp['Average'] for dp in write_iops_response['Datapoints']}
                    
                    total_iops_values = []
                    for timestamp in read_dict:
                        if timestamp in write_dict:
                            total_iops_values.append(read_dict[timestamp] + write_dict[timestamp])
                    
                    if total_iops_values:
                        metrics['iops-avg.1Hr/15-day period'] = round(sum(total_iops_values) / len(total_iops_values), 2)
                        
            except Exception as e:
                self.logger.warning(f"获取IOPS指标失败: {e}")
            
            # CPU指标
            cpu_response = self.cloudwatch_client.get_metric_statistics(
                Namespace='AWS/RDS',
                MetricName='CPUUtilization',
                Dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_identifier}],
                StartTime=start_time,
                EndTime=end_time,
                Period=3600,
                Statistics=['Average', 'Maximum', 'Minimum']
            )
            
            if cpu_response['Datapoints']:
                cpu_values = [dp['Average'] for dp in cpu_response['Datapoints']]
                metrics['cpu-avg.1Hr/15-day period'] = sum(cpu_values) / len(cpu_values)
                metrics['cpu-max.1Hr/15-day period'] = max(dp['Maximum'] for dp in cpu_response['Datapoints'])
                metrics['cpu-min.1Hr/15-day period'] = min(dp['Minimum'] for dp in cpu_response['Datapoints'])
            
        except Exception as e:
            self.logger.warning(f"获取CloudWatch指标失败: {e}")
        
        return metrics

    def get_all_mysql_pricing(self) -> None:
        """一次性获取RDS MySQL和Aurora MySQL定价"""
        if self.pricing_cache and self.aurora_pricing_cache:
            self.logger.info("MySQL定价缓存已存在，跳过重复获取")
            return
        
        try:
            self.logger.info(f"开始一次性获取区域 {self.region} 的所有MySQL实例定价...")
            
            # 分别获取两种引擎，但在一个方法中处理
            engines = ['MySQL', 'Aurora MySQL']
            rds_count = aurora_count = 0
            
            for engine in engines:
                filters = [
                    {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                    {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': engine},
                    {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()}
                ]
                
                if engine == 'MySQL':
                    filters.append({'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'})
                
                paginator = self.pricing_client.get_paginator('get_products')
                page_iterator = paginator.paginate(ServiceCode='AmazonRDS', Filters=filters)
                for page in page_iterator:
                    for price_item in page['PriceList']:
                        product = json.loads(price_item)
                        attributes = product.get('product', {}).get('attributes', {})
                        instance_type = attributes.get('instanceType')
                        
                        if not instance_type or not instance_type.startswith('db.'):
                            continue
                        
                        terms = product.get('terms', {})
                        reserved_terms = terms.get('Reserved', {})
                        
                        for term_key, term_data in reserved_terms.items():
                            term_attributes = term_data.get('termAttributes', {})
                            if (term_attributes.get('LeaseContractLength') == '1yr' and 
                                term_attributes.get('PurchaseOption') == 'No Upfront'):
                                
                                price_dimensions = term_data.get('priceDimensions', {})
                                for price_key, price_data in price_dimensions.items():
                                    description = price_data.get('description', '')
                                    
                                    # Aurora MySQL: 跳过IO-optimized
                                    if engine == 'Aurora MySQL' and 'IO-optimized' in description:
                                        continue
                                    
                                    price_per_unit = price_data.get('pricePerUnit', {})
                                    usd_price = price_per_unit.get('USD')
                                    if usd_price:
                                        hourly_rate = float(usd_price)
                                        if engine == 'MySQL':
                                            self.pricing_cache[instance_type] = hourly_rate
                                            rds_count += 1
                                        else:
                                            self.aurora_pricing_cache[instance_type] = hourly_rate
                                            aurora_count += 1
                                        break
                                break
            
            self.logger.info(f"RDS MySQL定价缓存完成，共缓存 {rds_count} 个实例类型的定价信息")
            self.logger.info(f"Aurora MySQL定价缓存完成，共缓存 {aurora_count} 个实例类型的定价信息")
            
        except Exception as e:
            self.logger.error(f"获取MySQL定价失败: {e}")
            self.logger.warning("MySQL定价缓存可能不完整，可能影响分析结果")

    def get_all_pg_pricing(self) -> None:
        """一次性获取RDS PostgreSQL和Aurora PostgreSQL定价"""
        if self.pg_pricing_cache and self.aurora_pg_pricing_cache:
            self.logger.info("PostgreSQL定价缓存已存在，跳过重复获取")
            return
        
        try:
            self.logger.info(f"开始一次性获取区域 {self.region} 的所有PostgreSQL实例定价...")
            
            engines = ['PostgreSQL', 'Aurora PostgreSQL']
            rds_count = aurora_count = 0
            
            for engine in engines:
                filters = [
                    {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                    {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': engine},
                    {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()}
                ]
                
                if engine == 'PostgreSQL':
                    filters.append({'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'})
                
                paginator = self.pricing_client.get_paginator('get_products')
                page_iterator = paginator.paginate(ServiceCode='AmazonRDS', Filters=filters)
                for page in page_iterator:
                    for price_item in page['PriceList']:
                        product = json.loads(price_item)
                        attributes = product.get('product', {}).get('attributes', {})
                        instance_type = attributes.get('instanceType')
                        
                        if not instance_type or not instance_type.startswith('db.'):
                            continue
                        
                        terms = product.get('terms', {})
                        reserved_terms = terms.get('Reserved', {})
                        
                        for term_key, term_data in reserved_terms.items():
                            term_attributes = term_data.get('termAttributes', {})
                            if (term_attributes.get('LeaseContractLength') == '1yr' and 
                                term_attributes.get('PurchaseOption') == 'No Upfront'):
                                
                                price_dimensions = term_data.get('priceDimensions', {})
                                for price_key, price_data in price_dimensions.items():
                                    description = price_data.get('description', '')
                                    
                                    # Aurora PostgreSQL: 跳过IO-optimized
                                    if engine == 'Aurora PostgreSQL' and 'IO-optimized' in description:
                                        continue
                                    
                                    price_per_unit = price_data.get('pricePerUnit', {})
                                    usd_price = price_per_unit.get('USD')
                                    if usd_price:
                                        hourly_rate = float(usd_price)
                                        if engine == 'PostgreSQL':
                                            self.pg_pricing_cache[instance_type] = hourly_rate
                                            rds_count += 1
                                        else:
                                            self.aurora_pg_pricing_cache[instance_type] = hourly_rate
                                            aurora_count += 1
                                        break
                                break
            
            self.logger.info(f"RDS PostgreSQL定价缓存完成，共缓存 {rds_count} 个实例类型的定价信息")
            self.logger.info(f"Aurora PostgreSQL定价缓存完成，共缓存 {aurora_count} 个实例类型的定价信息")
            
        except Exception as e:
            self.logger.error(f"获取PostgreSQL定价失败: {e}")
            self.logger.warning("PostgreSQL定价缓存可能不完整，可能影响分析结果")

    def get_all_storage_pricing(self) -> None:
        """获取所有存储类型定价（采用原始脚本的可靠方法）"""
        if self.storage_pricing_cache:
            self.logger.info("存储定价缓存已存在，跳过重复获取")
            return
        
        try:
            self.logger.info(f"开始获取区域 {self.region} 的所有存储类型定价...")
            
            # 支持的存储类型映射（与原始脚本一致）
            storage_type_mapping = {
                'General Purpose': 'gp2',
                'General Purpose-GP3': 'gp3', 
                'Provisioned IOPS': 'io1',
                'Provisioned IOPS-IO2': 'io2',
                'Magnetic': 'magnetic'
            }
            
            # 构建基础过滤器 - 获取所有RDS MySQL存储定价
            filters = [
                {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()},
                {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Database Storage'},
                {'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'},
                {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'MySQL'}
            ]
            
            # 使用分页获取所有存储产品
            paginator = self.pricing_client.get_paginator('get_products')
            page_iterator = paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=filters
            )
            
            # 初始化所有存储类型的定价信息
            for storage_type in storage_type_mapping.values():
                self.storage_pricing_cache[storage_type] = {
                    'storage_gb_month': 0,
                    'iops_month': 0,
                    'throughput_mbps_month': 0
                }
            
            processed_count = 0
            for page in page_iterator:
                for price_item in page['PriceList']:
                    product = json.loads(price_item)
                    attributes = product.get('product', {}).get('attributes', {})
                    
                    # 获取存储类型
                    volume_type = attributes.get('volumeType')
                    if not volume_type or volume_type not in storage_type_mapping:
                        continue
                    
                    storage_type = storage_type_mapping[volume_type]
                    
                    # 获取按需定价
                    terms = product.get('terms', {})
                    on_demand_terms = terms.get('OnDemand', {})
                    
                    for term_key, term_data in on_demand_terms.items():
                        price_dimensions = term_data.get('priceDimensions', {})
                        for price_key, price_data in price_dimensions.items():
                            price_per_unit = price_data.get('pricePerUnit', {})
                            usd_price = price_per_unit.get('USD')
                            unit = price_data.get('unit', '')
                            
                            if usd_price and float(usd_price) > 0:
                                price_value = float(usd_price)
                                
                                # 根据单位确定价格类型
                                if 'GB-Mo' in unit:
                                    self.storage_pricing_cache[storage_type]['storage_gb_month'] = price_value
                                elif 'IOPS-Mo' in unit:
                                    self.storage_pricing_cache[storage_type]['iops_month'] = price_value
                                elif 'MBps-Mo' in unit or 'MB/s-Mo' in unit:
                                    self.storage_pricing_cache[storage_type]['throughput_mbps_month'] = price_value
                                
                                processed_count += 1
            
            # 获取io1和io2的IOPS定价（使用专门的方法）
            self._get_iops_pricing_for_provisioned_storage()
            
            # 记录缓存结果
            for storage_type, pricing in self.storage_pricing_cache.items():
                if pricing['storage_gb_month'] > 0:
                    self.logger.info(f"存储类型 {storage_type} 定价: "
                                   f"存储=${pricing['storage_gb_month']}/GB/月, "
                                   f"IOPS=${pricing['iops_month']}/IOPS/月, "
                                   f"吞吐量=${pricing['throughput_mbps_month']}/MB/s/月")
            
            self.logger.info(f"存储定价缓存完成，共处理 {processed_count} 个定价项")
            
        except Exception as e:
            self.logger.error(f"获取存储定价失败: {e}")

    def _get_iops_pricing_for_provisioned_storage(self) -> None:
        """获取io1/io2/gp3存储类型的MySQL IOPS定价"""
        self.logger.info("开始获取io1/io2/gp3的MySQL IOPS定价...")
        
        mysql_results = {}
        general_results = {}
        
        try:
            paginator = self.pricing_client.get_paginator('get_products')
            for page in paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=[
                    {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Provisioned IOPS'},
                    {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()}
                ]
            ):
                for product_str in page['PriceList']:
                    product = json.loads(product_str)
                    attrs = product['product']['attributes']
                    usage_type = attrs.get('usagetype', '')
                    if 'Multi-AZ' in usage_type or 'Mirror' in usage_type:
                        continue
                    terms = product.get('terms', {}).get('OnDemand', {})
                    for term_data in terms.values():
                        for price_data in term_data.get('priceDimensions', {}).values():
                            unit = price_data.get('unit', '')
                            desc = price_data.get('description', '')
                            if 'IOPS' not in unit:
                                continue
                            price_per_unit = price_data.get('pricePerUnit', {})
                            if 'USD' not in price_per_unit:
                                continue
                            price = float(price_per_unit['USD'])
                            dl = desc.lower()
                            if 'io1' in dl:
                                storage_type = 'io1'
                            elif 'io2' in dl:
                                storage_type = 'io2'
                            elif 'gp3' in dl:
                                storage_type = 'gp3'
                            else:
                                continue
                            if 'MySQL' in desc:
                                if storage_type not in mysql_results or price < mysql_results[storage_type]:
                                    mysql_results[storage_type] = price
                            if storage_type not in general_results or price < general_results[storage_type]:
                                general_results[storage_type] = price
            
            for storage_type in ['io1', 'io2', 'gp3']:
                final_price = mysql_results.get(storage_type) or general_results.get(storage_type)
                if final_price:
                    if storage_type not in self.storage_pricing_cache:
                        self.storage_pricing_cache[storage_type] = {
                            'storage_gb_month': 0, 'iops_month': 0, 'throughput_mbps_month': 0
                        }
                    self.storage_pricing_cache[storage_type]['iops_month'] = final_price
                    self.logger.info(f"MySQL {storage_type} IOPS定价: ${final_price}/IOPS/月")
                
        except Exception as e:
            self.logger.error(f"获取MySQL IOPS定价失败: {e}")

    def get_all_pg_storage_pricing(self) -> None:
        """获取所有PostgreSQL存储类型定价"""
        if self.pg_storage_pricing_cache:
            self.logger.info("PostgreSQL存储定价缓存已存在，跳过重复获取")
            return
        
        try:
            self.logger.info(f"开始获取区域 {self.region} 的所有PostgreSQL存储类型定价...")
            
            storage_type_mapping = {
                'General Purpose': 'gp2',
                'General Purpose-GP3': 'gp3', 
                'Provisioned IOPS': 'io1',
                'Provisioned IOPS-IO2': 'io2',
                'Magnetic': 'magnetic'
            }
            
            filters = [
                {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()},
                {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Database Storage'},
                {'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'},
                {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'PostgreSQL'}
            ]
            
            paginator = self.pricing_client.get_paginator('get_products')
            page_iterator = paginator.paginate(ServiceCode='AmazonRDS', Filters=filters)
            
            for storage_type in storage_type_mapping.values():
                self.pg_storage_pricing_cache[storage_type] = {
                    'storage_gb_month': 0,
                    'iops_month': 0,
                    'throughput_mbps_month': 0
                }
            
            processed_count = 0
            for page in page_iterator:
                for price_item in page['PriceList']:
                    product = json.loads(price_item)
                    attributes = product.get('product', {}).get('attributes', {})
                    
                    volume_type = attributes.get('volumeType')
                    if not volume_type or volume_type not in storage_type_mapping:
                        continue
                    
                    storage_type = storage_type_mapping[volume_type]
                    
                    terms = product.get('terms', {})
                    on_demand_terms = terms.get('OnDemand', {})
                    
                    for term_key, term_data in on_demand_terms.items():
                        price_dimensions = term_data.get('priceDimensions', {})
                        for price_key, price_data in price_dimensions.items():
                            price_per_unit = price_data.get('pricePerUnit', {})
                            usd_price = price_per_unit.get('USD')
                            unit = price_data.get('unit', '')
                            
                            if usd_price and float(usd_price) > 0:
                                price_value = float(usd_price)
                                
                                if 'GB-Mo' in unit:
                                    self.pg_storage_pricing_cache[storage_type]['storage_gb_month'] = price_value
                                elif 'IOPS-Mo' in unit:
                                    self.pg_storage_pricing_cache[storage_type]['iops_month'] = price_value
                                elif 'MBps-Mo' in unit or 'MB/s-Mo' in unit:
                                    self.pg_storage_pricing_cache[storage_type]['throughput_mbps_month'] = price_value
                                
                                processed_count += 1
            
            self._get_iops_pricing_for_pg_provisioned_storage()
            
            for storage_type, pricing in self.pg_storage_pricing_cache.items():
                if pricing['storage_gb_month'] > 0:
                    self.logger.info(f"PG存储类型 {storage_type} 定价: "
                                   f"存储=${pricing['storage_gb_month']}/GB/月, "
                                   f"IOPS=${pricing['iops_month']}/IOPS/月, "
                                   f"吞吐量=${pricing['throughput_mbps_month']}/MB/s/月")
            
            self.logger.info(f"PostgreSQL存储定价缓存完成，共处理 {processed_count} 个定价项")
            
        except Exception as e:
            self.logger.error(f"获取PostgreSQL存储定价失败: {e}")

    def _get_iops_pricing_for_pg_provisioned_storage(self) -> None:
        """获取io1/io2/gp3存储类型的PostgreSQL IOPS定价"""
        self.logger.info("开始获取io1/io2/gp3的PostgreSQL IOPS定价...")
        
        pg_results = {}
        general_results = {}
        
        try:
            paginator = self.pricing_client.get_paginator('get_products')
            for page in paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=[
                    {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Provisioned IOPS'},
                    {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()}
                ]
            ):
                for product_str in page['PriceList']:
                    product = json.loads(product_str)
                    attrs = product['product']['attributes']
                    usage_type = attrs.get('usagetype', '')
                    if 'Multi-AZ' in usage_type or 'Mirror' in usage_type:
                        continue
                    terms = product.get('terms', {}).get('OnDemand', {})
                    for term_data in terms.values():
                        for price_data in term_data.get('priceDimensions', {}).values():
                            unit = price_data.get('unit', '')
                            desc = price_data.get('description', '')
                            if 'IOPS' not in unit:
                                continue
                            price_per_unit = price_data.get('pricePerUnit', {})
                            if 'USD' not in price_per_unit:
                                continue
                            price = float(price_per_unit['USD'])
                            dl = desc.lower()
                            if 'io1' in dl:
                                storage_type = 'io1'
                            elif 'io2' in dl:
                                storage_type = 'io2'
                            elif 'gp3' in dl:
                                storage_type = 'gp3'
                            else:
                                continue
                            if 'PostgreSQL' in desc:
                                if storage_type not in pg_results or price < pg_results[storage_type]:
                                    pg_results[storage_type] = price
                            if storage_type not in general_results or price < general_results[storage_type]:
                                general_results[storage_type] = price
            
            for storage_type in ['io1', 'io2', 'gp3']:
                final_price = pg_results.get(storage_type) or general_results.get(storage_type)
                if final_price:
                    if storage_type not in self.pg_storage_pricing_cache:
                        self.pg_storage_pricing_cache[storage_type] = {
                            'storage_gb_month': 0, 'iops_month': 0, 'throughput_mbps_month': 0
                        }
                    self.pg_storage_pricing_cache[storage_type]['iops_month'] = final_price
                    self.logger.info(f"PostgreSQL {storage_type} IOPS定价: ${final_price}/IOPS/月")
                
        except Exception as e:
            self.logger.error(f"获取PostgreSQL IOPS定价失败: {e}")

    def get_aurora_storage_pricing(self) -> Dict:
        """获取Aurora存储定价"""
        if self.aurora_storage_pricing_cache:
            self.logger.info("Aurora存储定价缓存已存在，跳过重复获取")
            return self.aurora_storage_pricing_cache
        
        try:
            self.logger.info(f"开始获取区域 {self.region} 的Aurora存储定价...")
            
            pricing_info = {
                'storage_gb_month': 0,
                'io_million_requests': 0
            }
            
            # 获取Aurora存储单价
            storage_filters = [
                {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()},
                {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Database Storage'},
                {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'Aurora MySQL'}
            ]
            
            storage_response = self.pricing_client.get_products(
                ServiceCode='AmazonRDS',
                Filters=storage_filters,
                MaxResults=10
            )
            
            if storage_response['PriceList']:
                product = json.loads(storage_response['PriceList'][0])
                terms = product.get('terms', {}).get('OnDemand', {})
                
                for term_data in terms.values():
                    for price_data in term_data.get('priceDimensions', {}).values():
                        usd_price = price_data.get('pricePerUnit', {}).get('USD')
                        if usd_price and 'GB-Mo' in price_data.get('unit', ''):
                            pricing_info['storage_gb_month'] = round(float(usd_price)/2.25,3)
                            self.logger.info(f"Aurora存储单价: ${round(float(usd_price)/2.25,3)}/GB/月")
                            break
                    else:
                        continue
                    break
            
            # 获取Aurora IO请求单价 - 使用实时API
            self.logger.info("通过API获取Aurora IO定价...")
            
            # 使用分页器获取Aurora产品
            paginator = self.pricing_client.get_paginator('get_products')
            
            for page in paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=[
                    {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'Aurora MySQL'},
                    {'Type': 'TERM_MATCH', 'Field': 'regionCode', 'Value': self.region}
                ]
            ):
                for product_str in page['PriceList']:
                    product = json.loads(product_str)
                    
                    # 获取定价信息
                    terms = product.get('terms', {}).get('OnDemand', {})
                    for term_data in terms.values():
                        for price_data in term_data.get('priceDimensions', {}).values():
                            unit = price_data.get('unit', '')
                            
                            # 查找IO单位的定价
                            if unit == 'IOs':
                                price_per_unit = price_data.get('pricePerUnit', {})
                                price_per_io = price_per_unit.get('USD', '0')
                                
                                if float(price_per_io) > 0:
                                    price_per_million = float(price_per_io) * 1000000
                                    pricing_info['io_million_requests'] = price_per_million
                                    self.logger.info(f"Aurora IO请求单价: ${price_per_million}/百万次请求")
                                    break
                        else:
                            continue
                        break
                    if pricing_info['io_million_requests'] > 0:
                        break
                if pricing_info['io_million_requests'] > 0:
                    break
            
            # 如果API未获取到IO定价，设置为0并记录错误
            if pricing_info['io_million_requests'] == 0:
                self.logger.error(f"未获取到Aurora IO定价")
            
            self.aurora_storage_pricing_cache = pricing_info
            self.logger.info(f"Aurora存储定价获取完成")
            return self.aurora_storage_pricing_cache
            
        except Exception as e:
            self.logger.error(f"获取Aurora存储定价失败: {e}")
            raise Exception(f"无法获取Aurora存储定价，请检查网络连接和AWS凭证: {e}")

    def get_aurora_pg_storage_pricing(self) -> Dict:
        """获取Aurora PostgreSQL存储定价"""
        if self.aurora_pg_storage_pricing_cache:
            self.logger.info("Aurora PostgreSQL存储定价缓存已存在，跳过重复获取")
            return self.aurora_pg_storage_pricing_cache
        
        try:
            self.logger.info(f"开始获取区域 {self.region} 的Aurora PostgreSQL存储定价...")
            
            pricing_info = {
                'storage_gb_month': 0,
                'io_million_requests': 0
            }
            
            # 获取Aurora PostgreSQL存储单价
            storage_filters = [
                {'Type': 'TERM_MATCH', 'Field': 'servicecode', 'Value': 'AmazonRDS'},
                {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': self._get_pricing_location()},
                {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'Database Storage'},
                {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'Aurora PostgreSQL'}
            ]
            
            storage_response = self.pricing_client.get_products(
                ServiceCode='AmazonRDS',
                Filters=storage_filters,
                MaxResults=10
            )
            
            if storage_response['PriceList']:
                product = json.loads(storage_response['PriceList'][0])
                terms = product.get('terms', {}).get('OnDemand', {})
                
                for term_data in terms.values():
                    for price_data in term_data.get('priceDimensions', {}).values():
                        usd_price = price_data.get('pricePerUnit', {}).get('USD')
                        if usd_price and 'GB-Mo' in price_data.get('unit', ''):
                            pricing_info['storage_gb_month'] = round(float(usd_price)/2.25, 3)
                            self.logger.info(f"Aurora PostgreSQL存储单价: ${round(float(usd_price)/2.25, 3)}/GB/月")
                            break
                    else:
                        continue
                    break
            
            # 获取Aurora PostgreSQL IO请求单价
            self.logger.info("通过API获取Aurora PostgreSQL IO定价...")
            
            paginator = self.pricing_client.get_paginator('get_products')
            
            for page in paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=[
                    {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': 'Aurora PostgreSQL'},
                    {'Type': 'TERM_MATCH', 'Field': 'regionCode', 'Value': self.region}
                ]
            ):
                for product_str in page['PriceList']:
                    product = json.loads(product_str)
                    
                    terms = product.get('terms', {}).get('OnDemand', {})
                    for term_data in terms.values():
                        for price_data in term_data.get('priceDimensions', {}).values():
                            unit = price_data.get('unit', '')
                            
                            if unit == 'IOs':
                                price_per_unit = price_data.get('pricePerUnit', {})
                                price_per_io = price_per_unit.get('USD', '0')
                                
                                if float(price_per_io) > 0:
                                    price_per_million = float(price_per_io) * 1000000
                                    pricing_info['io_million_requests'] = price_per_million
                                    self.logger.info(f"Aurora PostgreSQL IO请求单价: ${price_per_million}/百万次请求")
                                    break
                        else:
                            continue
                        break
                    if pricing_info['io_million_requests'] > 0:
                        break
                if pricing_info['io_million_requests'] > 0:
                    break
            
            if pricing_info['io_million_requests'] == 0:
                self.logger.error(f"未获取到Aurora PostgreSQL IO定价")
            
            self.aurora_pg_storage_pricing_cache = pricing_info
            self.logger.info(f"Aurora PostgreSQL存储定价获取完成")
            return self.aurora_pg_storage_pricing_cache
            
        except Exception as e:
            self.logger.error(f"获取Aurora PostgreSQL存储定价失败: {e}")
            raise Exception(f"无法获取Aurora PostgreSQL存储定价，请检查网络连接和AWS凭证: {e}")

    def get_all_rds_mysql_instances(self) -> List[Dict]:
        """获取所有RDS MySQL实例"""
        if self.enable_optimization:
            return self._get_instances_optimized()
        else:
            return self._get_instances_original()

    def _get_instances_original(self) -> List[Dict]:
        """原始模式获取实例"""
        instances = []
        account_id = self.get_account_id()
        
        try:
            # 获取所有Aurora集群信息
            aurora_clusters = {}
            try:
                paginator = self.rds_client.get_paginator('describe_db_clusters')
                for page in paginator.paginate():
                    for cluster in page['DBClusters']:
                        if cluster.get('Engine') == 'aurora-mysql':
                            aurora_clusters[cluster['DBClusterIdentifier']] = cluster
            except Exception as e:
                self.logger.warning(f"获取Aurora集群信息失败: {e}")
            
            # 获取所有RDS实例
            all_db_instances = []
            paginator = self.rds_client.get_paginator('describe_db_instances')
            
            for page in paginator.paginate():
                for db_instance in page['DBInstances']:
                    if (db_instance.get('Engine') == 'mysql' and
                        db_instance.get('DBInstanceStatus') in ['available', 'storage-optimization', 'modifying', 'backing-up'] and
                        self._should_process_instance(db_instance['DBInstanceClass'])):
                        all_db_instances.append(db_instance)
            
            # 处理每个实例
            for db_instance in all_db_instances:
                instance_data = self._process_single_instance_complete(db_instance, all_db_instances, aurora_clusters, account_id, engine_type='mysql')
                if instance_data:
                    instances.append(instance_data)
            
            self.logger.info(f"找到 {len(instances)} 个符合条件的RDS MySQL实例")
            return instances
            
        except Exception as e:
            self.logger.error(f"获取RDS实例失败: {e}")
            return []

    def _get_instances_optimized(self) -> List[Dict]:
        """优化模式获取实例"""
        instances = []
        account_id = self.get_account_id()
        
        try:
            # 获取所有Aurora集群信息
            aurora_clusters = {}
            try:
                paginator = self.rds_client.get_paginator('describe_db_clusters')
                for page in paginator.paginate():
                    for cluster in page['DBClusters']:
                        if cluster.get('Engine') == 'aurora-mysql':
                            aurora_clusters[cluster['DBClusterIdentifier']] = cluster
            except Exception as e:
                self.logger.warning(f"获取Aurora集群信息失败: {e}")
            
            # 获取所有实例
            all_instances = []
            paginator = self.rds_client.get_paginator('describe_db_instances')
            
            for page in paginator.paginate():
                for db_instance in page['DBInstances']:
                    if (db_instance.get('Engine') == 'mysql' and
                        db_instance.get('DBInstanceStatus') in ['available', 'storage-optimization', 'modifying', 'backing-up'] and
                        self._should_process_instance(db_instance['DBInstanceClass'])):
                        all_instances.append(db_instance)
            
            self.logger.info(f"开始并发处理 {len(all_instances)} 个实例，使用 {self.max_workers} 个线程")
            
            # 并发处理实例
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_instance = {
                    executor.submit(self._process_single_instance_complete, db_instance, all_instances, aurora_clusters, account_id, 'mysql'): db_instance
                    for db_instance in all_instances
                }
                
                for future in as_completed(future_to_instance):
                    try:
                        result = future.result()
                        if result:
                            instances.append(result)
                    except Exception as e:
                        db_instance = future_to_instance[future]
                        self.logger.error(f"处理实例 {db_instance.get('DBInstanceIdentifier', 'unknown')} 失败: {e}")
            
            self.logger.info(f"并发处理完成，成功处理 {len(instances)} 个实例")
            return instances
            
        except Exception as e:
            self.logger.error(f"优化模式获取实例失败: {e}")
            return []

    def get_all_rds_pg_instances(self) -> List[Dict]:
        """获取所有RDS PostgreSQL实例"""
        if self.enable_optimization:
            return self._get_pg_instances_optimized()
        else:
            return self._get_pg_instances_original()

    def _get_pg_instances_original(self) -> List[Dict]:
        """原始模式获取PostgreSQL实例"""
        instances = []
        account_id = self.get_account_id()
        
        try:
            # 获取所有Aurora PostgreSQL集群信息
            aurora_clusters = {}
            try:
                paginator = self.rds_client.get_paginator('describe_db_clusters')
                for page in paginator.paginate():
                    for cluster in page['DBClusters']:
                        if cluster.get('Engine') == 'aurora-postgresql':
                            aurora_clusters[cluster['DBClusterIdentifier']] = cluster
            except Exception as e:
                self.logger.warning(f"获取Aurora PostgreSQL集群信息失败: {e}")
            
            all_db_instances = []
            paginator = self.rds_client.get_paginator('describe_db_instances')
            
            for page in paginator.paginate():
                for db_instance in page['DBInstances']:
                    if (db_instance.get('Engine') == 'postgres' and
                        db_instance.get('DBInstanceStatus') in ['available', 'storage-optimization', 'modifying'] and
                        self._should_process_instance(db_instance['DBInstanceClass'])):
                        all_db_instances.append(db_instance)
            
            for db_instance in all_db_instances:
                instance_data = self._process_single_instance_complete(db_instance, all_db_instances, aurora_clusters, account_id, engine_type='postgresql')
                if instance_data:
                    instances.append(instance_data)
            
            self.logger.info(f"找到 {len(instances)} 个符合条件的RDS PostgreSQL实例")
            return instances
            
        except Exception as e:
            self.logger.error(f"获取RDS PostgreSQL实例失败: {e}")
            return []

    def _get_pg_instances_optimized(self) -> List[Dict]:
        """优化模式获取PostgreSQL实例"""
        instances = []
        account_id = self.get_account_id()
        
        try:
            aurora_clusters = {}
            try:
                paginator = self.rds_client.get_paginator('describe_db_clusters')
                for page in paginator.paginate():
                    for cluster in page['DBClusters']:
                        if cluster.get('Engine') == 'aurora-postgresql':
                            aurora_clusters[cluster['DBClusterIdentifier']] = cluster
            except Exception as e:
                self.logger.warning(f"获取Aurora PostgreSQL集群信息失败: {e}")
            
            all_instances = []
            paginator = self.rds_client.get_paginator('describe_db_instances')
            
            for page in paginator.paginate():
                for db_instance in page['DBInstances']:
                    if (db_instance.get('Engine') == 'postgres' and
                        db_instance.get('DBInstanceStatus') in ['available', 'storage-optimization', 'modifying'] and
                        self._should_process_instance(db_instance['DBInstanceClass'])):
                        all_instances.append(db_instance)
            
            self.logger.info(f"开始并发处理 {len(all_instances)} 个PostgreSQL实例，使用 {self.max_workers} 个线程")
            
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_instance = {
                    executor.submit(self._process_single_instance_complete, db_instance, all_instances, aurora_clusters, account_id, engine_type='postgresql'): db_instance
                    for db_instance in all_instances
                }
                
                for future in as_completed(future_to_instance):
                    try:
                        result = future.result()
                        if result:
                            instances.append(result)
                    except Exception as e:
                        db_instance = future_to_instance[future]
                        self.logger.error(f"处理PostgreSQL实例 {db_instance.get('DBInstanceIdentifier', 'unknown')} 失败: {e}")
            
            self.logger.info(f"并发处理完成，成功处理 {len(instances)} 个PostgreSQL实例")
            return instances
            
        except Exception as e:
            self.logger.error(f"优化模式获取PostgreSQL实例失败: {e}")
            return []

    def _process_single_instance_complete(self, db_instance: Dict, all_db_instances: List[Dict], 
                                        aurora_clusters: Dict, account_id: str, engine_type: str = 'mysql') -> Optional[Dict]:
        """完整处理单个实例"""
        try:
            instance_id = db_instance['DBInstanceIdentifier']
            instance_class = db_instance['DBInstanceClass']
            engine_version = db_instance.get('EngineVersion', 'unknown')
            
            # 确定架构类型
            architecture = self._determine_architecture_with_taz(db_instance, all_db_instances, aurora_clusters)
            
            # 获取存储信息
            storage_info = self.get_storage_info(instance_id, db_instance)
            
            # 获取CloudWatch指标
            cloudwatch_metrics = self.get_cloudwatch_metrics(instance_id)
            
            # 获取source_db_instance_identifier - 与原始脚本逻辑一致
            source_db_identifier = db_instance.get('ReadReplicaSourceDBInstanceIdentifier', '')
            
            # 处理集群级Read Replica
            if not source_db_identifier and db_instance.get('ReadReplicaSourceDBClusterIdentifier'):
                # 集群级Read Replica，需要找到源集群的Writer节点
                source_cluster = db_instance['ReadReplicaSourceDBClusterIdentifier']
                try:
                    cluster_response = self.rds_client.describe_db_clusters(DBClusterIdentifier=source_cluster)
                    cluster = cluster_response['DBClusters'][0]
                    for member in cluster.get('DBClusterMembers', []):
                        if member.get('IsClusterWriter', False):
                            source_db_identifier = member['DBInstanceIdentifier']
                            break
                except Exception as e:
                    self.logger.warning(f"获取源集群Writer失败: {e}")
            
            # 对于TAZ集群，通过API获取集群信息来确定writer节点
            if architecture == 'taz_cluster':
                cluster_id = db_instance.get('DBClusterIdentifier')
                if cluster_id:
                    try:
                        response = self.rds_client.describe_db_clusters(DBClusterIdentifier=cluster_id)
                        cluster = response['DBClusters'][0]
                        members = cluster.get('DBClusterMembers', [])
                        
                        # 找到Writer节点
                        writer_instance = None
                        for member in members:
                            if member.get('IsClusterWriter'):
                                writer_instance = member['DBInstanceIdentifier']
                                break
                        
                        if writer_instance and writer_instance != instance_id:
                            # 当前实例是Reader，设置source为Writer
                            source_db_identifier = writer_instance
                            self.logger.info(f"TAZ集群Reader节点 {instance_id} 的source_db_instance_identifier设置为Writer节点: {writer_instance}")
                        else:
                            # 当前实例是Writer
                            source_db_identifier = 'N/A'
                            
                    except Exception as e:
                        self.logger.warning(f"获取TAZ集群Writer失败: {e}")
                        source_db_identifier = 'N/A'
            
            # 对于RDS MySQL Read Replica，保持原有逻辑
            elif not source_db_identifier:
                source_db_identifier = 'N/A'
            
            return {
                'db_instance_identifier': instance_id,
                'account_id': account_id,
                'region': self.region,
                'engine_type': engine_type,
                'engine_version': engine_version,
                'architecture': architecture,
                'instance_class': instance_class,
                'source_db_instance_identifier': source_db_identifier,
                'read_replica_source_db_cluster_identifier': db_instance.get('ReadReplicaSourceDBClusterIdentifier', ''),
                **storage_info,
                **cloudwatch_metrics
            }
            
        except Exception as e:
            self.logger.error(f"处理实例失败: {e}")
            return None

    def _determine_architecture_with_taz(self, db_instance: Dict, all_instances: List[Dict], aurora_clusters: Dict) -> str:
        """
        优化的架构判断 - 仅通过API，正确识别TAZ和主从关系
        """
        
        # 1. Read Replica优先，基于自身的MultiAZ状态
        if db_instance.get('ReadReplicaSourceDBInstanceIdentifier'):
            return 'multi_az' if db_instance.get('MultiAZ', False) else 'read_replica'
        
        # 2. 集群架构判断
        cluster_id = db_instance.get('DBClusterIdentifier')
        if cluster_id and cluster_id != 'N/A':
            try:
                response = self.rds_client.describe_db_clusters(DBClusterIdentifier=cluster_id)
                cluster = response['DBClusters'][0]
                members = cluster.get('DBClusterMembers', [])
                
                # TAZ: 3成员 + 1Writer/2Reader + 跨3AZ
                if (len(members) == 3 and 
                    len([m for m in members if m.get('IsClusterWriter')]) == 1 and
                    self._is_cross_three_az(members, all_instances)):
                    self.logger.info(f"实例 {db_instance['DBInstanceIdentifier']} 属于TAZ集群")
                    return 'taz_cluster'
                    
                return 'multi_az' if cluster.get('MultiAZ') else 'single_az'
            except Exception as e:
                self.logger.warning(f"获取集群信息失败: {e}")
        
        # 3. 实例级MultiAZ
        return 'multi_az' if db_instance.get('MultiAZ') else 'single_az'

    def _is_cross_three_az(self, members: List[Dict], all_instances: List[Dict]) -> bool:
        """判断集群成员是否跨3个AZ"""
        azs = set()
        instance_map = {inst['DBInstanceIdentifier']: inst for inst in all_instances}
        for member in members:
            member_id = member.get('DBInstanceIdentifier', '')
            inst = instance_map.get(member_id)
            if inst:
                az = inst.get('AvailabilityZone', '')
                if az:
                    azs.add(az)
        return len(azs) >= 3

    def _get_cluster_info(self, db_instance: Dict, aurora_clusters: Dict) -> Dict:
        """获取集群信息"""
        cluster_id = db_instance.get('DBClusterIdentifier')
        if cluster_id and cluster_id in aurora_clusters:
            return aurora_clusters[cluster_id]
        return {}

    def analyze_instances(self) -> List[Dict]:
        """主分析方法"""
        self.logger.info(f"开始分析 - 模式: {'优化' if self.enable_optimization else '标准'}, 引擎: {self.engine}, mode: {self.mode}")
        
        all_results = []
        
        # MySQL 分析
        if self.engine in ('mysql', 'all'):
            self.logger.info("=" * 60)
            self.logger.info("开始分析 RDS MySQL 实例")
            self.logger.info("=" * 60)
            
            self.get_all_mysql_pricing()
            self.get_all_storage_pricing()
            if self.mode in ('rds-to-aurora', 'all'):
                self.get_aurora_storage_pricing()
            
            instances = self.get_all_rds_mysql_instances()
            
            rds_mysql_clusters = self.identify_rds_clusters(instances)
            self.logger.info(f"识别到 {len(rds_mysql_clusters)} 个RDS MySQL集群")
            
            for instance in instances:
                instance_id = instance['db_instance_identifier']
                for primary, replicas in rds_mysql_clusters.items():
                    replica_ids = [inst['db_instance_identifier'] if isinstance(inst, dict) else inst for inst in replicas]
                    if instance_id in replica_ids and len(replica_ids) >= 2 and instance_id != primary:
                        instance['source_db_instance_identifier'] = primary
                        self.logger.info(f"RDS MySQL集群Reader节点 {instance_id} 的source_db_instance_identifier设置为Primary节点: {primary}")
            
            cluster_aurora_storage_costs = {}
            if self.mode in ('rds-to-aurora', 'all'):
                cluster_aurora_storage_costs = self._calculate_cluster_aurora_storage_costs(instances, engine_type='mysql')
            
            for instance in instances:
                self.logger.info(f"分析MySQL实例: {instance['db_instance_identifier']}")
                result = self._analyze_single_instance(instance, cluster_aurora_storage_costs, rds_mysql_clusters, instances)
                if result:
                    all_results.append(result)
        
        # PostgreSQL 分析
        if self.engine in ('postgresql', 'all'):
            self.logger.info("=" * 60)
            self.logger.info("开始分析 RDS PostgreSQL 实例")
            self.logger.info("=" * 60)
            
            self.get_all_pg_pricing()
            self.get_all_pg_storage_pricing()
            if self.mode in ('rds-to-aurora', 'all'):
                self.get_aurora_pg_storage_pricing()
            
            pg_instances = self.get_all_rds_pg_instances()
            
            rds_pg_clusters = self.identify_rds_clusters(pg_instances)
            self.logger.info(f"识别到 {len(rds_pg_clusters)} 个RDS PostgreSQL集群")
            
            for instance in pg_instances:
                instance_id = instance['db_instance_identifier']
                for primary, replicas in rds_pg_clusters.items():
                    replica_ids = [inst['db_instance_identifier'] if isinstance(inst, dict) else inst for inst in replicas]
                    if instance_id in replica_ids and len(replica_ids) >= 2 and instance_id != primary:
                        instance['source_db_instance_identifier'] = primary
                        self.logger.info(f"RDS PostgreSQL集群Reader节点 {instance_id} 的source_db_instance_identifier设置为Primary节点: {primary}")
            
            cluster_aurora_pg_storage_costs = {}
            if self.mode in ('rds-to-aurora', 'all'):
                cluster_aurora_pg_storage_costs = self._calculate_cluster_aurora_storage_costs(pg_instances, engine_type='postgresql')
            
            for instance in pg_instances:
                self.logger.info(f"分析PostgreSQL实例: {instance['db_instance_identifier']}")
                result = self._analyze_single_instance(instance, cluster_aurora_pg_storage_costs, rds_pg_clusters, pg_instances)
                if result:
                    all_results.append(result)
        
        return all_results

    def identify_rds_clusters(self, instances: List[Dict]) -> Dict[str, List[str]]:
        """识别RDS集群（MySQL和PostgreSQL通用）"""
        clusters = {}
        
        for instance in instances:
            instance_id = instance['db_instance_identifier']
            source_id = instance.get('source_db_instance_identifier', '')
            
            if source_id:
                # 这是一个Read Replica
                if source_id not in clusters:
                    clusters[source_id] = [source_id]
                if instance_id not in clusters[source_id]:
                    clusters[source_id].append(instance_id)
                self.logger.info(f"实例 {instance_id} 是 {source_id} 的Read Replica")
        
        return clusters

    def _calculate_cluster_aurora_storage_costs(self, instances: List[Dict], engine_type: str = 'mysql') -> Dict[str, float]:
        """计算集群Aurora存储成本 - 简化逻辑"""
        cluster_costs = {}
        if engine_type == 'postgresql':
            storage_price = self.aurora_pg_storage_pricing_cache.get('storage_gb_month', 0.1)
        else:
            storage_price = self.aurora_storage_pricing_cache.get('storage_gb_month', 0.1)
        
        for instance in instances:
            instance_id = instance['db_instance_identifier']
            source_id = instance.get('source_db_instance_identifier', '')
            
            # 只有Primary/Writer节点计算存储成本
            if not source_id or source_id in ['N/A', '']:
                used_storage = instance.get('used_storage_gb', 0)
                if used_storage <= 0:
                    used_storage = instance.get('allocated_storage_gb', 0)
                
                storage_cost = used_storage * storage_price
                cluster_costs[instance_id] = round(storage_cost, 2)
                self.logger.info(f"Primary/Writer节点 {instance_id} Aurora存储成本: ${storage_cost:.2f} (存储: {used_storage}GB)")
            else:
                # Read Replica节点存储成本为0
                cluster_costs[instance_id] = 0
                self.logger.info(f"Read Replica节点 {instance_id} Aurora存储成本为0")
        
        return cluster_costs

    def _analyze_single_instance(self, instance: Dict, cluster_aurora_storage_costs: Dict, rds_clusters: Dict = None, all_instances: List[Dict] = None) -> Dict:
        """分析单个实例"""
        instance_id = instance['db_instance_identifier']
        instance_class = instance['instance_class']
        architecture = instance['architecture']
        engine_type = instance.get('engine_type', 'mysql')
        
        # 根据引擎类型选择定价缓存
        if engine_type == 'postgresql':
            pricing_cache = self.pg_pricing_cache
            aurora_pricing_cache = self.aurora_pg_pricing_cache
            storage_pricing_cache = self.pg_storage_pricing_cache
            aurora_storage_pricing_cache = self.aurora_pg_storage_pricing_cache
        else:
            pricing_cache = self.pricing_cache
            aurora_pricing_cache = self.aurora_pricing_cache
            storage_pricing_cache = self.storage_pricing_cache
            aurora_storage_pricing_cache = self.aurora_storage_pricing_cache
        
        # 获取RDS定价
        rds_hourly_rate = pricing_cache.get(instance_class, 0)
        rds_adjusted_rate = self._adjust_pricing_for_architecture(rds_hourly_rate, architecture)
        rds_instance_mrr = round(rds_adjusted_rate * 730, 2) if rds_adjusted_rate else 0
        
        # 计算存储成本
        storage_cost_info = self._calculate_storage_cost(instance, architecture, engine_type)
        rds_total_mrr = rds_instance_mrr + storage_cost_info['total_storage_cost_per_month_usd']
        
        # 基础结果
        result = {
            'account_id': instance['account_id'],
            'region': instance['region'],
            'engine_type': engine_type,
            'cluster_primary_instance': self._get_cluster_primary_instance(instance, rds_clusters),
            'engine_version': instance.get('engine_version', ''),
            'db_instance_identifier': instance_id,
            'source_db_instance_identifier': instance.get('source_db_instance_identifier', '') or '',
            'architecture': architecture,
            'rds_instance_class': instance_class,
            'rds_hourly_rate_usd': rds_adjusted_rate if rds_adjusted_rate else 'NA',
            'rds_instance_mrr_usd': rds_instance_mrr,
            'storage_type': instance.get('storage_type', ''),
            'allocated_storage_gb': instance.get('allocated_storage_gb', 0),
            'used_storage_gb': instance.get('used_storage_gb', 0),
            'iops': instance.get('iops', 0),
            'storage_throughput_mbps': instance.get('storage_throughput', 0),
            'iops-avg.1Hr/15-day period': instance.get('iops-avg.1Hr/15-day period', 0),
            'cpu-max.1Hr/15-day period': instance.get('cpu-max.1Hr/15-day period', 0),
            'cpu-min.1Hr/15-day period': instance.get('cpu-min.1Hr/15-day period', 0),
            'cpu-avg.1Hr/15-day period': instance.get('cpu-avg.1Hr/15-day period', 0),
            **storage_cost_info,
            'rds_total_mrr_usd': round(rds_total_mrr, 2),
        }
        
        # Aurora 存储字段仅在 rds-to-aurora 或 all 模式下填充（IO-Optimized 模式）
        if self.mode in ('rds-to-aurora', 'all'):
            std_storage_price = aurora_storage_pricing_cache.get('storage_gb_month', 0.1)
            io_opt_storage_price = round(std_storage_price * 2.25, 4)
            std_storage_cost = cluster_aurora_storage_costs.get(instance_id, 0)
            # aurora_allocate_storage_gb 基于 Standard 成本反推容量
            aurora_storage_gb = std_storage_cost / std_storage_price if std_storage_cost > 0 and std_storage_price > 0 else 0
            io_opt_storage_cost = round(aurora_storage_gb * io_opt_storage_price, 2)
            
            result['aurora_allocate_storage_gb'] = round(aurora_storage_gb, 2)
            result['aurora_io_opt_storage_price_per_gb_month_usd'] = io_opt_storage_price if std_storage_cost > 0 else 0
            result['aurora_io_opt_total_storage_cost_per_month_usd'] = io_opt_storage_cost
        
        # Aurora分析（仅 rds-to-aurora 或 all 模式）
        if self.mode in ('rds-to-aurora', 'all'):
            for generation in self.aurora_generations:
                aurora_result = self._analyze_aurora_conversion(instance, generation, cluster_aurora_storage_costs, rds_clusters, all_instances)
                result.update(aurora_result)
        
        # RDS替换分析（仅 rds-replacement 或 all 模式）
        if self.mode in ('rds-replacement', 'all'):
            for generation in self.rds_replacement_generations:
                rds_replacement_result = self._analyze_rds_replacement(instance, generation, storage_cost_info)
                result.update(rds_replacement_result)
        
        return result

    def _get_cluster_primary_instance(self, instance: Dict, rds_clusters: Dict = None) -> str:
        """获取集群主实例（与原始脚本逻辑一致）"""
        instance_id = instance['db_instance_identifier']
        
        # 检查是否是RDS MySQL集群的primary节点
        if rds_clusters and instance_id in rds_clusters:
            return instance_id
        
        # 检查是否是RDS MySQL集群的read replica
        if rds_clusters:
            for primary_id, cluster_instances in rds_clusters.items():
                cluster_instance_ids = [inst['db_instance_identifier'] if isinstance(inst, dict) else inst for inst in cluster_instances]
                if instance_id in cluster_instance_ids and instance_id != primary_id:
                    return primary_id
        
        # 检查TAZ集群 - 通过API获取集群信息
        architecture = instance.get('architecture', '')
        if architecture == 'taz_cluster':
            source_db = instance.get('source_db_instance_identifier')
            if source_db and source_db != 'N/A' and source_db.strip():
                return source_db
            else:
                return instance_id  # Writer节点
        
        # 检查source_db_instance_identifier
        source_db = instance.get('source_db_instance_identifier')
        if source_db and source_db != 'N/A' and source_db.strip():
            return source_db
        
        # 默认返回实例自身
        return instance_id

    def _adjust_pricing_for_architecture(self, hourly_rate: float, architecture: str) -> float:
        """根据架构调整定价（与原始脚本一致）"""
        if architecture in ['multi_az']:
            self.logger.debug(f"实例是{architecture}，RDS成本乘以2")
            return hourly_rate * 2
        elif architecture == 'taz_cluster':
            self.logger.debug(f"实例是TAZ集群成员，使用单价")
            return hourly_rate
        return hourly_rate

    def _calculate_storage_cost(self, instance: Dict, architecture: str, engine_type: str = 'mysql') -> Dict:
        """计算存储成本"""
        storage_type = instance.get('storage_type', 'gp2')
        allocated_gb = instance.get('allocated_storage_gb', 0)
        iops = instance.get('iops', 0)
        
        # 根据引擎类型选择存储定价缓存
        if engine_type == 'postgresql':
            storage_pricing = self.pg_storage_pricing_cache.get(storage_type, {})
        else:
            storage_pricing = self.storage_pricing_cache.get(storage_type, {})
        storage_price_per_gb = storage_pricing.get('storage_gb_month', 0.115)
        iops_price_per_iops = storage_pricing.get('iops_month', 0)
        
        # 计算存储容量成本
        storage_cost = allocated_gb * storage_price_per_gb
        
        # 计算IOPS成本
        # gp3: 只对超出基线的部分收费 (< 400GB: 3000 free, >= 400GB: 12000 free)
        if storage_type == 'gp3':
            baseline_iops = 12000 if allocated_gb >= 400 else 3000
            billable_iops = max(0, iops - baseline_iops)
            iops_cost = billable_iops * iops_price_per_iops
        else:
            iops_cost = iops * iops_price_per_iops
        
        total_cost = storage_cost + iops_cost
        
        # Multi-AZ架构存储成本乘以2
        if architecture == 'multi_az':
            storage_cost *= 2
            iops_cost *= 2
            total_cost *= 2
        
        return {
            'storage_price_per_gb_month_usd': storage_price_per_gb,
            'iops_price_per_iops_month_usd': iops_price_per_iops,
            'storage_cost_per_month_usd': round(storage_cost, 2),
            'iops_cost_per_month_usd': round(iops_cost, 2),
            'total_storage_cost_per_month_usd': round(total_cost, 2)
        }

    def _analyze_aurora_conversion(self, instance: Dict, generation: str, cluster_aurora_storage_costs: Dict, rds_clusters: Dict = None, all_instances: List[Dict] = None) -> Dict:
        """分析Aurora转换
        
        Aurora节点映射规则（集群级别）：
        - Primary SAZ，无replica           → 1W = 1节点
        - Primary SAZ + N replica          → 1W + N×1R = 1+N节点
        - Primary MAZ，无replica           → 1W + 1R = 2节点
        - Primary MAZ + 1 replica(SAZ)     → 1W + 1R = 2节点
        - Primary MAZ + 1 replica(MAZ/TAZ) → 1W + 2R = 3节点
        - Primary MAZ + N replica(N≥2)     → 1W + N×1R = 1+N节点（不管replica架构）
        
        单实例级别：每个实例只算自己映射的Aurora节点数，集群总数=所有实例之和。
        """
        instance_id = instance['db_instance_identifier']
        instance_class = instance['instance_class']
        architecture = instance['architecture']
        engine_type = instance.get('engine_type', 'mysql')
        source_id = instance.get('source_db_instance_identifier', '') or ''
        
        # 根据引擎类型选择Aurora定价缓存
        if engine_type == 'postgresql':
            aurora_pricing_cache = self.aurora_pg_pricing_cache
        else:
            aurora_pricing_cache = self.aurora_pricing_cache
        
        # 映射到Aurora实例
        aurora_instance_class = self._map_to_aurora_instance(instance_class, generation)
        
        # 获取Aurora定价
        aurora_hourly_rate = aurora_pricing_cache.get(aurora_instance_class, 0)
        
        # 计算该实例映射的Aurora节点数
        is_replica = bool(source_id and source_id not in ['N/A', ''])
        
        if is_replica:
            aurora_instance_count = 1  # default: every replica = 1 Aurora reader
            # 特殊情况：非 TAZ primary + 有且仅有 1 个 MAZ/TAZ replica → replica 映射 2 个节点
            source_id_val = instance.get('source_db_instance_identifier', '') or ''
            if source_id_val and source_id_val not in ['N/A', ''] and rds_clusters and source_id_val in rds_clusters:
                primary_arch = None
                if all_instances:
                    for inst in all_instances:
                        if inst.get('db_instance_identifier') == source_id_val:
                            primary_arch = inst.get('architecture', '')
                            break
                if primary_arch != 'taz_cluster':
                    members = rds_clusters[source_id_val]
                    rep_ids = [m if isinstance(m, str) else m.get('db_instance_identifier', '')
                               for m in members if (m if isinstance(m, str) else m.get('db_instance_identifier', '')) != source_id_val]
                    if len(rep_ids) == 1 and architecture in ['multi_az', 'taz_cluster']:
                        aurora_instance_count = 2
        else:
            # Primary
            replica_count = 0
            replica_ids = []
            if rds_clusters and instance_id in rds_clusters:
                members = rds_clusters[instance_id]
                replica_ids = [m if isinstance(m, str) else m.get('db_instance_identifier', '') 
                              for m in members if (m if isinstance(m, str) else m.get('db_instance_identifier', '')) != instance_id]
                replica_count = len(replica_ids)
            
            if architecture == 'taz_cluster':
                # TAZ primary: 始终 1W，不管有多少 replica
                aurora_instance_count = 1
            elif architecture != 'multi_az':
                aurora_instance_count = 1  # SAZ primary
            elif replica_count == 0:
                aurora_instance_count = 2  # MAZ no replica: 1W + 1R
            elif replica_count == 1:
                replica_id = replica_ids[0]
                replica_arch = 'single_az'
                if all_instances:
                    for inst in all_instances:
                        if inst.get('db_instance_identifier') == replica_id:
                            replica_arch = inst.get('architecture', 'single_az')
                            break
                if replica_arch in ['multi_az', 'taz_cluster']:
                    aurora_instance_count = 1  # MAZ + 1 MAZ/TAZ replica: primary=1W, replica行算2
                else:
                    aurora_instance_count = 1  # MAZ + 1 SAZ replica: primary=1W
            else:
                aurora_instance_count = 1  # MAZ + N>=2: 1W
        
        aurora_instance_mrr = round(aurora_hourly_rate * 730 * aurora_instance_count, 2) if aurora_hourly_rate else 0
        
        # 获取存储成本（IO-Optimized模式下IO已包含在实例费×1.3和存储×2.25中）
        storage_cost = cluster_aurora_storage_costs.get(instance_id, 0)
        
        # 如果该代次在此region无定价，所有成本标记为NA
        if not aurora_hourly_rate:
            return {
                f'aurora_{generation}_instance_class': aurora_instance_class,
                f'aurora_{generation}_instance_count': aurora_instance_count,
                f'aurora_{generation}_io_opt_hourly_rate_usd': 'NA',
                f'aurora_{generation}_io_opt_instance_mrr_usd': 'NA',
                f'aurora_{generation}_io_opt_total_mrr_usd': 'NA'
            }
        
        # IO-Optimized模式计算：实例费用*1.3，存储*2.25（无IO费用）
        io_opt_hourly_rate = round(aurora_hourly_rate * 1.3, 6)
        io_opt_instance_mrr = round(io_opt_hourly_rate * 730 * aurora_instance_count, 2)
        io_opt_total_cost = io_opt_instance_mrr + storage_cost * 2.25
        
        return {
            f'aurora_{generation}_instance_class': aurora_instance_class,
            f'aurora_{generation}_instance_count': aurora_instance_count,
            f'aurora_{generation}_io_opt_hourly_rate_usd': io_opt_hourly_rate,
            f'aurora_{generation}_io_opt_instance_mrr_usd': io_opt_instance_mrr,
            f'aurora_{generation}_io_opt_total_mrr_usd': round(io_opt_total_cost, 2)
        }

    def _analyze_rds_replacement(self, instance: Dict, generation: str, storage_cost_info: Dict) -> Dict:
        """分析RDS替换"""
        instance_class = instance['instance_class']
        architecture = instance['architecture']
        engine_type = instance.get('engine_type', 'mysql')
        
        # 根据引擎类型选择定价缓存
        if engine_type == 'postgresql':
            pricing_cache = self.pg_pricing_cache
        else:
            pricing_cache = self.pricing_cache
        
        # 映射到RDS替换实例
        replacement_instance_class = self._map_to_rds_replacement_instance(instance_class, generation)
        
        if not replacement_instance_class:
            return {
                f'rds_replacement_{generation}_instance_class': 'NA',
                f'rds_replacement_{generation}_hourly_rate_usd': 'NA',
                f'rds_replacement_{generation}_instance_mrr_usd': 'NA',
                f'rds_replacement_{generation}_storage_mrr_usd': 'NA',
                f'rds_replacement_{generation}_total_mrr_usd': 'NA'
            }
        
        # 获取RDS替换定价
        replacement_hourly_rate = pricing_cache.get(replacement_instance_class, 0)
        
        # 如果该代次在此region无定价，所有成本标记为NA
        if not replacement_hourly_rate:
            return {
                f'rds_replacement_{generation}_instance_class': replacement_instance_class,
                f'rds_replacement_{generation}_hourly_rate_usd': 'NA',
                f'rds_replacement_{generation}_instance_mrr_usd': 'NA',
                f'rds_replacement_{generation}_storage_mrr_usd': 'NA',
                f'rds_replacement_{generation}_total_mrr_usd': 'NA'
            }
        
        replacement_adjusted_rate = self._adjust_pricing_for_architecture(replacement_hourly_rate, architecture)
        replacement_instance_mrr = round(replacement_adjusted_rate * 730, 2) if replacement_adjusted_rate else 0
        
        # 存储成本
        storage_mrr = storage_cost_info['total_storage_cost_per_month_usd']
        total_mrr = replacement_instance_mrr + storage_mrr
        
        return {
            f'rds_replacement_{generation}_instance_class': replacement_instance_class,
            f'rds_replacement_{generation}_hourly_rate_usd': replacement_adjusted_rate if replacement_adjusted_rate else 'NA',
            f'rds_replacement_{generation}_instance_mrr_usd': replacement_instance_mrr,
            f'rds_replacement_{generation}_storage_mrr_usd': storage_mrr,
            f'rds_replacement_{generation}_total_mrr_usd': round(total_mrr, 2)
        }

    def _map_to_aurora_instance(self, rds_instance_class: str, aurora_generation: str) -> str:
        """映射到Aurora实例"""
        instance_type = rds_instance_class.replace('db.', '')
        parts = instance_type.split('.')
        
        if len(parts) >= 2:
            family = parts[0]
            size = parts[1]
            # Aurora 最小规格是 large，T 系列的 micro/small/medium 统一映射到 large
            if family.startswith('t') or size in ('micro', 'small', 'medium'):
                size = 'large'
            return f"db.{aurora_generation}.{size}"
        else:
            return f"db.{aurora_generation}.large"

    def _map_to_rds_replacement_instance(self, rds_instance_class: str, replacement_generation: str) -> Optional[str]:
        """映射到RDS替换实例"""
        instance_type = rds_instance_class.replace('db.', '')
        parts = instance_type.split('.')
        
        if len(parts) >= 2:
            instance_family = parts[0]
            size = parts[1]
            
            # M系列只能映射到m6g/m7g/m8g
            if instance_family.startswith('m') and replacement_generation in ['m6g', 'm7g', 'm8g']:
                return f"db.{replacement_generation}.{size}"
            # R系列只能映射到r6g/r7g/r8g
            elif instance_family.startswith('r') and replacement_generation in ['r6g', 'r7g', 'r8g']:
                return f"db.{replacement_generation}.{size}"
            # C系列映射到m6g/m7g/m8g
            elif instance_family.startswith('c') and replacement_generation in ['m6g', 'm7g', 'm8g']:
                return f"db.{replacement_generation}.{size}"
            # T系列映射到m6g/m7g/m8g（T系列通常是小规格，映射到M系列）
            elif instance_family.startswith('t') and replacement_generation in ['m6g', 'm7g', 'm8g']:
                return f"db.{replacement_generation}.{size}"
        
        return None

    # ====================================================================
    # Aurora Provisioned Graviton 换代评估（只算 instance cost）
    # ====================================================================

    def get_all_aurora_provisioned_instances(self, engine_type: str = 'mysql') -> List[Dict]:
        """获取所有 Aurora MySQL/PG provisioned 实例（排除 Serverless）"""
        instances = []
        account_id = self.get_account_id()
        
        engine_filter = 'aurora-mysql' if engine_type == 'mysql' else 'aurora-postgresql'
        
        try:
            # 获取所有 Aurora 集群
            clusters = {}
            paginator = self.rds_client.get_paginator('describe_db_clusters')
            for page in paginator.paginate():
                for cluster in page['DBClusters']:
                    if cluster.get('Engine') == engine_filter:
                        clusters[cluster['DBClusterIdentifier']] = cluster
            
            # 获取所有实例
            paginator = self.rds_client.get_paginator('describe_db_instances')
            for page in paginator.paginate():
                for db_instance in page['DBInstances']:
                    # 仅处理 Aurora 引擎、available 状态、provisioned 实例
                    if (db_instance.get('Engine') == engine_filter and
                        db_instance.get('DBInstanceStatus') in ['available', 'modifying', 'backing-up'] and
                        db_instance.get('DBInstanceClass', '').startswith('db.')):
                        
                        instance_class = db_instance['DBInstanceClass']
                        # 排除 Serverless（db.serverless）
                        if 'serverless' in instance_class.lower():
                            continue
                        
                        cluster_id = db_instance.get('DBClusterIdentifier', '')
                        cluster_info = clusters.get(cluster_id, {})
                        
                        # 判断该实例是 Writer 还是 Reader
                        is_writer = False
                        for member in cluster_info.get('DBClusterMembers', []):
                            if member.get('DBInstanceIdentifier') == db_instance['DBInstanceIdentifier']:
                                is_writer = member.get('IsClusterWriter', False)
                                break
                        
                        instances.append({
                            'db_instance_identifier': db_instance['DBInstanceIdentifier'],
                            'db_cluster_identifier': cluster_id,
                            'account_id': account_id,
                            'region': self.region,
                            'engine_type': engine_type,
                            'engine': engine_filter,
                            'engine_version': db_instance.get('EngineVersion', ''),
                            'instance_class': instance_class,
                            'is_writer': is_writer,
                            'availability_zone': db_instance.get('AvailabilityZone', ''),
                            'multi_az': db_instance.get('MultiAZ', False),
                        })
            
            self.logger.info(f"找到 {len(instances)} 个 Aurora {'MySQL' if engine_type == 'mysql' else 'PostgreSQL'} provisioned 实例")
            return instances
            
        except Exception as e:
            self.logger.error(f"获取Aurora provisioned实例失败: {e}")
            return []

    def _get_aurora_instance_generation(self, instance_class: str) -> str:
        """提取 Aurora 实例代次，例如 db.r6g.large -> r6g, db.r5.xlarge -> r5"""
        parts = instance_class.replace('db.', '').split('.')
        return parts[0] if parts else ''

    def _get_aurora_instance_size(self, instance_class: str) -> str:
        """提取 Aurora 实例大小，例如 db.r6g.large -> large"""
        parts = instance_class.replace('db.', '').split('.')
        return parts[1] if len(parts) >= 2 else 'large'

    def _is_already_latest_graviton(self, generation: str) -> bool:
        """判断当前代次是否已是最新 Graviton（r8g）"""
        return generation in ['r8g']

    def _get_graviton_generation_order(self) -> List[str]:
        """返回 Graviton 代次顺序（从旧到新），用于判断当前代次是否 >= 目标代次"""
        return ['r5', 'r5b', 'r5d', 'x2g', 'r6g', 'r6i', 'r6gd', 'r7g', 'r7i', 'r8g']

    def _is_current_gen_gte_target(self, current_generation: str, target_generation: str) -> bool:
        """
        判断当前代次是否 >= 目标代次
        例如: r7g >= r6g → True, r7g >= r7g → True, r5 >= r7g → False
        """
        order = self._get_graviton_generation_order()
        # 规范化比较：r6g/r6i/r6gd 都算作 r6g 级别，r7g/r7i 都算 r7g 级别
        gen_level = {
            'r5': 0, 'r5b': 0, 'r5d': 0, 'x2g': 0,
            'r6g': 1, 'r6i': 1, 'r6gd': 1,
            'r7g': 2, 'r7i': 2,
            'r8g': 3,
        }
        current_level = gen_level.get(current_generation, -1)
        target_level = gen_level.get(target_generation, 99)
        return current_level >= target_level

    def analyze_aurora_graviton_upgrade(self) -> List[Dict]:
        """
        分析 Aurora provisioned 实例的 Graviton 换代评估
        使用 IO-Optimized 定价（Standard RI × 1.3），只算 instance cost
        所有 R 系列（除 T）都计算 r6g/r7g/r8g 成本，当前代次 >= 目标代次的列显示 '--'
        """
        all_results = []
        
        # MySQL
        if self.engine in ('mysql', 'all'):
            self.logger.info("=" * 60)
            self.logger.info("开始 Aurora MySQL Provisioned Graviton 换代评估 (IO-Optimized)")
            self.logger.info("=" * 60)
            
            # 确保 Aurora MySQL 定价已加载
            if not self.aurora_pricing_cache:
                self.get_all_mysql_pricing()
            
            mysql_instances = self.get_all_aurora_provisioned_instances('mysql')
            for instance in mysql_instances:
                result = self._analyze_aurora_graviton_single(instance, self.aurora_pricing_cache)
                if result:
                    all_results.append(result)
        
        # PostgreSQL
        if self.engine in ('postgresql', 'all'):
            self.logger.info("=" * 60)
            self.logger.info("开始 Aurora PostgreSQL Provisioned Graviton 换代评估 (IO-Optimized)")
            self.logger.info("=" * 60)
            
            # 确保 Aurora PostgreSQL 定价已加载
            if not self.aurora_pg_pricing_cache:
                self.get_all_pg_pricing()
            
            pg_instances = self.get_all_aurora_provisioned_instances('postgresql')
            for instance in pg_instances:
                result = self._analyze_aurora_graviton_single(instance, self.aurora_pg_pricing_cache)
                if result:
                    all_results.append(result)
        
        self.logger.info(f"Aurora Graviton换代评估完成，共 {len(all_results)} 个实例")
        return all_results

    def _analyze_aurora_graviton_single(self, instance: Dict, pricing_cache: Dict) -> Optional[Dict]:
        """
        分析单个 Aurora 实例的 Graviton 换代成本（IO-Optimized 模式）
        
        规则：
        - IO-Optimized instance cost = Standard RI hourly × 1.3 × 730
        - 所有 R 系列都计算 r6g/r7g/r8g 三列成本
        - 如果当前代次 >= 目标代次，该列显示 '--'
        - 排除 T 系列
        """
        instance_class = instance['instance_class']
        current_generation = self._get_aurora_instance_generation(instance_class)
        size = self._get_aurora_instance_size(instance_class)
        
        # 排除 T 系列
        if current_generation.startswith('t'):
            return None
        
        # 获取当前实例 IO-Optimized 定价 = Standard RI × 1.3
        current_std_hourly = pricing_cache.get(instance_class, 0)
        current_hourly_rate = round(current_std_hourly * 1.3, 6) if current_std_hourly else 0
        current_instance_mrr = round(current_hourly_rate * 730, 2) if current_hourly_rate else 0
        
        result = {
            'account_id': instance['account_id'],
            'region': instance['region'],
            'engine_type': instance['engine_type'],
            'engine': instance['engine'],
            'engine_version': instance['engine_version'],
            'db_cluster_identifier': instance['db_cluster_identifier'],
            'db_instance_identifier': instance['db_instance_identifier'],
            'role': 'Writer' if instance['is_writer'] else 'Reader',
            'current_instance_class': instance_class,
            'current_generation': current_generation,
            'current_io_opt_hourly_rate_usd': current_hourly_rate if current_hourly_rate else 'NA',
            'current_io_opt_instance_mrr_usd': current_instance_mrr if current_hourly_rate else 'NA',
        }
        
        # 计算每个目标代次的 IO-Optimized 成本
        for target_gen in self.aurora_graviton_upgrade_targets:
            target_instance_class = f"db.{target_gen}.{size}"
            
            # 当前代次 >= 目标代次，显示 '--'
            if self._is_current_gen_gte_target(current_generation, target_gen):
                result[f'upgrade_{target_gen}_instance_class'] = '--'
                result[f'upgrade_{target_gen}_io_opt_hourly_rate_usd'] = '--'
                result[f'upgrade_{target_gen}_io_opt_instance_mrr_usd'] = '--'
                result[f'upgrade_{target_gen}_saving_mrr_usd'] = '--'
                result[f'upgrade_{target_gen}_saving_pct'] = '--'
            else:
                target_std_hourly = pricing_cache.get(target_instance_class, 0)
                target_hourly_rate = round(target_std_hourly * 1.3, 6) if target_std_hourly else 0
                
                if target_hourly_rate and current_hourly_rate:
                    target_instance_mrr = round(target_hourly_rate * 730, 2)
                    saving_mrr = round(current_instance_mrr - target_instance_mrr, 2)
                    saving_pct = round((current_instance_mrr - target_instance_mrr) / current_instance_mrr * 100, 1) if current_instance_mrr > 0 else 0
                    
                    result[f'upgrade_{target_gen}_instance_class'] = target_instance_class
                    result[f'upgrade_{target_gen}_io_opt_hourly_rate_usd'] = target_hourly_rate
                    result[f'upgrade_{target_gen}_io_opt_instance_mrr_usd'] = target_instance_mrr
                    result[f'upgrade_{target_gen}_saving_mrr_usd'] = saving_mrr
                    result[f'upgrade_{target_gen}_saving_pct'] = saving_pct
                else:
                    # 该 region 无此代次定价
                    result[f'upgrade_{target_gen}_instance_class'] = target_instance_class
                    result[f'upgrade_{target_gen}_io_opt_hourly_rate_usd'] = 'NA'
                    result[f'upgrade_{target_gen}_io_opt_instance_mrr_usd'] = 'NA'
                    result[f'upgrade_{target_gen}_saving_mrr_usd'] = 'NA'
                    result[f'upgrade_{target_gen}_saving_pct'] = 'NA'
        
        return result

    def generate_aurora_graviton_cluster_summary(self, results: List[Dict]) -> List[Dict]:
        """生成 Aurora Graviton 换代的集群级汇总（IO-Optimized 模式）"""
        clusters = {}
        
        for result in results:
            cluster_key = result.get('db_cluster_identifier', result.get('db_instance_identifier'))
            
            if cluster_key not in clusters:
                clusters[cluster_key] = {
                    'account_id': result.get('account_id'),
                    'region': result.get('region'),
                    'engine_type': result.get('engine_type'),
                    'engine_version': result.get('engine_version'),
                    'db_cluster_identifier': cluster_key,
                    'instance_count': 0,
                    'current_generation': result.get('current_generation'),
                    'current_io_opt_total_instance_mrr_usd': 0,
                }
                for target_gen in self.aurora_graviton_upgrade_targets:
                    clusters[cluster_key][f'upgrade_{target_gen}_io_opt_total_mrr_usd'] = 0
                    clusters[cluster_key][f'upgrade_{target_gen}_total_saving_mrr_usd'] = 0
            
            summary = clusters[cluster_key]
            summary['instance_count'] += 1
            
            current_mrr = result.get('current_io_opt_instance_mrr_usd', 0)
            if isinstance(current_mrr, (int, float)):
                summary['current_io_opt_total_instance_mrr_usd'] += current_mrr
            
            for target_gen in self.aurora_graviton_upgrade_targets:
                target_mrr = result.get(f'upgrade_{target_gen}_io_opt_instance_mrr_usd', 0)
                saving_mrr = result.get(f'upgrade_{target_gen}_saving_mrr_usd', 0)
                # 跳过 '--' 值（当前代次 >= 目标代次）
                if isinstance(target_mrr, (int, float)):
                    summary[f'upgrade_{target_gen}_io_opt_total_mrr_usd'] += target_mrr
                if isinstance(saving_mrr, (int, float)):
                    summary[f'upgrade_{target_gen}_total_saving_mrr_usd'] += saving_mrr
        
        # 计算集群级 saving_pct
        summary_list = []
        for summary in clusters.values():
            current_total = summary['current_io_opt_total_instance_mrr_usd']
            for target_gen in self.aurora_graviton_upgrade_targets:
                target_total = summary[f'upgrade_{target_gen}_io_opt_total_mrr_usd']
                if current_total > 0 and isinstance(target_total, (int, float)) and target_total > 0:
                    summary[f'upgrade_{target_gen}_saving_pct'] = round(
                        (current_total - target_total) / current_total * 100, 1)
                else:
                    summary[f'upgrade_{target_gen}_saving_pct'] = '--'
            
            # 四舍五入
            for key, value in summary.items():
                if key.endswith('_usd') and isinstance(value, (int, float)):
                    summary[key] = round(value, 2)
            summary_list.append(summary)
        
        return summary_list

    def export_to_excel(self, results: List[Dict], output_file: str) -> None:
        """导出到Excel，多Sheet + 冻结 + 格式化 + 集群超链接"""
        if not results and not self.aurora_provisioned_results:
            self.logger.warning("没有数据可导出")
            return
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
            from openpyxl.utils import get_column_letter
            import openpyxl as openpyxl_lib
            
            mysql_results = [r for r in results if r.get('engine_type') == 'mysql']
            pg_results = [r for r in results if r.get('engine_type') == 'postgresql']
            
            excel_file = output_file.replace('.csv', '.xlsx')
            wb = openpyxl_lib.Workbook()
            wb.remove(wb.active)  # remove default sheet
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
            link_font = Font(color="0563C1", underline="single")
            dollar_fmt = '$#,##0.00'
            pct_fmt = '0.0%'
            
            def _write_summary_sheet(ws, cluster_summary, detail_sheet_name, engine_label):
                """Write cluster summary with hyperlinks to detail sheet"""
                # Define columns for summary (conditional on mode)
                summary_cols = [
                    ('Account', 'account_id', None),
                    ('Region', 'region', None),
                    ('Cluster Primary', 'cluster_primary_instance', None),
                    ('Instances', 'instance_count', None),
                    ('RDS Current MRR', 'rds_total_mrr_usd', dollar_fmt),
                ]
                if self.mode in ('rds-to-aurora', 'all'):
                    summary_cols += [
                        ('Aurora r6g IO-Opt MRR', 'aurora_r6g_io_opt_total_mrr_usd', dollar_fmt),
                        ('Aurora r7g IO-Opt MRR', 'aurora_r7g_io_opt_total_mrr_usd', dollar_fmt),
                        ('Aurora r8g IO-Opt MRR', 'aurora_r8g_io_opt_total_mrr_usd', dollar_fmt),
                    ]
                if self.mode in ('rds-replacement', 'all'):
                    summary_cols += [
                        ('RDS r6g/m6g MRR', 'replacement_r6g_m6g_total_mrr_usd', dollar_fmt),
                        ('RDS r7g/m7g MRR', 'replacement_r7g_m7g_total_mrr_usd', dollar_fmt),
                        ('RDS r8g/m8g MRR', 'replacement_r8g_m8g_total_mrr_usd', dollar_fmt),
                    ]
                
                # Write headers
                for ci, (col_name, _, _) in enumerate(summary_cols, 1):
                    cell = ws.cell(row=1, column=ci, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data with hyperlinks
                for ri, row_data in enumerate(cluster_summary, 2):
                    for ci, (_, field, fmt) in enumerate(summary_cols, 1):
                        val = row_data.get(field, '')
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if fmt:
                            cell.number_format = fmt
                        # Add hyperlink on cluster name to detail sheet
                        if field == 'cluster_primary_instance' and val:
                            cell.font = link_font
                            cell.hyperlink = f"#{detail_sheet_name}!A1"
                
                # Freeze top row
                ws.freeze_panes = 'A2'
                # Auto-width
                for ci in range(1, len(summary_cols) + 1):
                    ws.column_dimensions[get_column_letter(ci)].width = 18
            
            def _write_detail_sheet(ws, detail_results):
                """Write instance detail with formatting"""
                if not detail_results:
                    return
                df = pd.DataFrame(detail_results)
                
                # Write headers
                for ci, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=ci, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
                # Write data
                for ri, row in enumerate(df.itertuples(index=False), 2):
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        col_name = df.columns[ci - 1]
                        # Format dollar columns
                        if '_mrr_' in col_name or '_cost_' in col_name or 'hourly_rate' in col_name or 'price_per' in col_name:
                            if isinstance(val, (int, float)):
                                cell.number_format = dollar_fmt
                
                # Freeze top row + first 3 columns
                ws.freeze_panes = 'D2'
                # Auto-width for first few columns
                for ci in range(1, min(6, len(df.columns) + 1)):
                    ws.column_dimensions[get_column_letter(ci)].width = 22
            
            # Create sheets
            if mysql_results:
                ws_summary = wb.create_sheet('MySQL集群汇总')
                ws_detail = wb.create_sheet('MySQL详细数据')
                mysql_cluster = self.generate_cluster_summary(mysql_results)
                _write_summary_sheet(ws_summary, mysql_cluster, 'MySQL详细数据', 'MySQL')
                _write_detail_sheet(ws_detail, mysql_results)
            
            if pg_results:
                ws_pg_summary = wb.create_sheet('PG集群汇总')
                ws_pg_detail = wb.create_sheet('PG详细数据')
                pg_cluster = self.generate_cluster_summary(pg_results)
                _write_summary_sheet(ws_pg_summary, pg_cluster, 'PG详细数据', 'PostgreSQL')
                _write_detail_sheet(ws_pg_detail, pg_results)
            
            # Aurora Provisioned Graviton 换代评估 Sheets
            if self.aurora_provisioned_results:
                aurora_grav_mysql = [r for r in self.aurora_provisioned_results if r.get('engine_type') == 'mysql']
                aurora_grav_pg = [r for r in self.aurora_provisioned_results if r.get('engine_type') == 'postgresql']
                
                if aurora_grav_mysql:
                    ws_agm_summary = wb.create_sheet('Aurora MySQL Graviton汇总')
                    ws_agm_detail = wb.create_sheet('Aurora MySQL Graviton详细')
                    agm_cluster_summary = self.generate_aurora_graviton_cluster_summary(aurora_grav_mysql)
                    self._write_aurora_graviton_summary_sheet(ws_agm_summary, agm_cluster_summary)
                    _write_detail_sheet(ws_agm_detail, aurora_grav_mysql)
                
                if aurora_grav_pg:
                    ws_agp_summary = wb.create_sheet('Aurora PG Graviton汇总')
                    ws_agp_detail = wb.create_sheet('Aurora PG Graviton详细')
                    agp_cluster_summary = self.generate_aurora_graviton_cluster_summary(aurora_grav_pg)
                    self._write_aurora_graviton_summary_sheet(ws_agp_summary, agp_cluster_summary)
                    _write_detail_sheet(ws_agp_detail, aurora_grav_pg)
            
            wb.save(excel_file)
            self.logger.info(f"Excel已导出: {excel_file}")
            
        except Exception as e:
            self.logger.error(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()

    def _write_aurora_graviton_summary_sheet(self, ws, cluster_summary: List[Dict]) -> None:
        """写入 Aurora Graviton 换代集群汇总 Sheet (IO-Optimized)"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
        dollar_fmt = '$#,##0.00'
        pct_fmt = '0.0"%"'
        
        # 定义列
        summary_cols = [
            ('Account', 'account_id', None),
            ('Region', 'region', None),
            ('Engine', 'engine_type', None),
            ('Engine Version', 'engine_version', None),
            ('Cluster', 'db_cluster_identifier', None),
            ('Instances', 'instance_count', None),
            ('Current Gen', 'current_generation', None),
            ('Current IO-Opt MRR', 'current_io_opt_total_instance_mrr_usd', dollar_fmt),
        ]
        
        for target_gen in self.aurora_graviton_upgrade_targets:
            summary_cols.append((f'{target_gen} IO-Opt MRR', f'upgrade_{target_gen}_io_opt_total_mrr_usd', dollar_fmt))
            summary_cols.append((f'{target_gen} Saving MRR', f'upgrade_{target_gen}_total_saving_mrr_usd', dollar_fmt))
            summary_cols.append((f'{target_gen} Saving %', f'upgrade_{target_gen}_saving_pct', pct_fmt))
        
        # 写入表头
        for ci, (col_name, _, _) in enumerate(summary_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for ri, row_data in enumerate(cluster_summary, 2):
            for ci, (_, field, fmt) in enumerate(summary_cols, 1):
                val = row_data.get(field, '')
                cell = ws.cell(row=ri, column=ci, value=val)
                if fmt and isinstance(val, (int, float)):
                    cell.number_format = fmt
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        for ci in range(1, len(summary_cols) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 20

    def generate_cluster_summary(self, results: List[Dict]) -> List[Dict]:
        """生成集群成本汇总（与原始脚本完全一致）"""
        clusters = {}
        
        for result in results:
            # 确定集群标识：优先使用source_db_instance_identifier的Primary节点，否则使用实例自身
            source_db = result.get('source_db_instance_identifier', 'N/A')
            if source_db != 'N/A' and source_db is not None and source_db.strip():
                cluster_key = source_db  # 使用Primary节点作为集群标识
            else:
                cluster_key = result.get('db_instance_identifier', 'unknown')  # Primary节点使用自身
            
            rds_instance_class = result.get('rds_instance_class', '')
            
            # 判断实例系列
            instance_family = rds_instance_class.replace('db.', '').split('.')[0] if rds_instance_class else ''
            series_type = 'R系列' if instance_family.startswith('r') else 'M/C系列'
            
            if cluster_key not in clusters:
                base_fields = {
                    'account_id': result.get('account_id'),
                    'region': result.get('region'),
                    'cluster_primary_instance': cluster_key,
                    'instance_count': 0,
                    'rds_total_mrr_usd': 0,
                }
                if self.mode in ('rds-to-aurora', 'all'):
                    base_fields['aurora_r6g_io_opt_total_mrr_usd'] = 0
                    base_fields['aurora_r7g_io_opt_total_mrr_usd'] = 0
                    base_fields['aurora_r8g_io_opt_total_mrr_usd'] = 0
                if self.mode in ('rds-replacement', 'all'):
                    base_fields['replacement_r6g_m6g_total_mrr_usd'] = 0
                    base_fields['replacement_r7g_m7g_total_mrr_usd'] = 0
                    base_fields['replacement_r8g_m8g_total_mrr_usd'] = 0
                
                clusters[cluster_key] = base_fields
            
            summary = clusters[cluster_key]
            
            # 使用Aurora节点数作为instance_count（rds-to-aurora模式），否则用1
            aurora_nodes = 1  # default
            if self.mode in ('rds-to-aurora', 'all'):
                for gen in self.aurora_generations:
                    nc = result.get(f'aurora_{gen}_instance_count', None)
                    if nc and isinstance(nc, (int, float)):
                        aurora_nodes = int(nc)
                        break
            summary['instance_count'] += aurora_nodes
            
            # 汇总成本字段
            summary['rds_total_mrr_usd'] += self._safe_float(result.get('rds_total_mrr_usd', 0))
            
            if self.mode in ('rds-to-aurora', 'all'):
                for field in ['aurora_r6g_io_opt_total_mrr_usd', 'aurora_r7g_io_opt_total_mrr_usd', 'aurora_r8g_io_opt_total_mrr_usd']:
                    value = result.get(field, 0)
                    if isinstance(value, (int, float)):
                        summary[field] += value
            
            if self.mode in ('rds-replacement', 'all'):
                # 合并RDS替换成本字段
                r6g_m6g_cost = (self._safe_float(result.get('rds_replacement_r6g_total_mrr_usd', 0)) + 
                               self._safe_float(result.get('rds_replacement_m6g_total_mrr_usd', 0)))
                r7g_m7g_cost = (self._safe_float(result.get('rds_replacement_r7g_total_mrr_usd', 0)) + 
                               self._safe_float(result.get('rds_replacement_m7g_total_mrr_usd', 0)))
                r8g_m8g_cost = (self._safe_float(result.get('rds_replacement_r8g_total_mrr_usd', 0)) + 
                               self._safe_float(result.get('rds_replacement_m8g_total_mrr_usd', 0)))
                
                summary['replacement_r6g_m6g_total_mrr_usd'] += r6g_m6g_cost
                summary['replacement_r7g_m7g_total_mrr_usd'] += r7g_m7g_cost
                summary['replacement_r8g_m8g_total_mrr_usd'] += r8g_m8g_cost
        
        # 转换为列表并四舍五入
        summary_list = []
        for summary in clusters.values():
            for key, value in summary.items():
                if key.endswith('_usd') and isinstance(value, (int, float)):
                    summary[key] = round(value, 2)
            summary_list.append(summary)
        
        return summary_list

    def _safe_float(self, value) -> float:
        """安全转换为float"""
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def run_analysis(self, output_file: str = None) -> None:
        """运行分析，根据 self.mode 决定执行哪些评估"""
        start_time = time.time()
        
        try:
            results = []
            
            # RDS → Aurora 迁移评估 或 RDS 机型替换（两者共用实例扫描，通过输出字段区分）
            if self.mode in ('rds-to-aurora', 'rds-replacement', 'all'):
                results = self.analyze_instances()
            
            # Aurora Provisioned Graviton 换代评估
            if self.mode in ('aurora-replacement', 'all'):
                self.aurora_provisioned_results = self.analyze_aurora_graviton_upgrade()
            
            if not output_file:
                mode_label = self.mode if self.mode != 'all' else ('optimized' if self.enable_optimization else 'standard')
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"rds_aurora_analysis_{mode_label}_{self.region}_{timestamp}.xlsx"
            
            self.export_to_excel(results, output_file)
            
            end_time = time.time()
            self.logger.info(f"分析完成，总耗时: {end_time - start_time:.2f}秒")
            
            # 打印总结信息
            self._print_summary(results, output_file)
            
        except Exception as e:
            self.logger.error(f"分析失败: {e}")
            raise

    def _print_summary(self, results: List[Dict], output_file: str) -> None:
        """打印分析总结"""
        mysql_count = len([r for r in results if r.get('engine_type') == 'mysql'])
        pg_count = len([r for r in results if r.get('engine_type') == 'postgresql'])
        aurora_grav_count = len(self.aurora_provisioned_results)
        aurora_mysql_grav = len([r for r in self.aurora_provisioned_results if r.get('engine_type') == 'mysql'])
        aurora_pg_grav = len([r for r in self.aurora_provisioned_results if r.get('engine_type') == 'postgresql'])
        
        print(f"\n分析完成！结果已保存到 {output_file}")
        print(f"RDS实例: {len(results)} 个 (MySQL: {mysql_count}, PostgreSQL: {pg_count})")
        print(f"Aurora Provisioned Graviton换代评估: {aurora_grav_count} 个 (MySQL: {aurora_mysql_grav}, PostgreSQL: {aurora_pg_grav})")
        print("\nRDS成本包含：")
        print("- 实例MRR：RDS实例的月度费用（不含存储）")
        print("- 存储MRR：RDS存储的月度费用")
        print("- 总MRR：实例MRR + 存储MRR")
        print("Aurora转换成本包含：")
        print("- 实例MRR：Aurora实例的月度费用（不含存储）")
        print("- 存储MRR：Aurora存储的月度费用（基于Primary节点的used_storage_gb）")
        print("- 总MRR：实例MRR + 存储MRR")
        print("RDS替换成本包含：")
        print("- 实例MRR：RDS替换实例的月度费用（不含存储）")
        print("- 存储MRR：RDS替换存储的月度费用")
        print("- 总MRR：实例MRR + 存储MRR")
        print("Aurora Graviton换代评估：")
        print("- 只计算 instance cost（RI 1yr No Upfront）")
        print("- 对比当前代次 → r7g/r8g 的实例月费节省")


def main():
    parser = argparse.ArgumentParser(
        description='RDS/Aurora 多代实例定价分析工具 — 支持 MySQL 和 PostgreSQL',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 全量分析（RDS→Aurora + RDS机型替换 + Aurora Graviton换代）
  python %(prog)s ap-southeast-1

  # 仅 RDS → Aurora 迁移评估
  python %(prog)s ap-southeast-1 --mode rds-to-aurora

  # 仅 RDS 机型替换评估（留在 RDS，换 Graviton）
  python %(prog)s ap-southeast-1 --mode rds-replacement

  # 仅 Aurora provisioned 实例 Graviton 换代评估
  python %(prog)s ap-southeast-1 --mode aurora-replacement

  # 只分析 MySQL 的 Aurora 换代
  python %(prog)s ap-southeast-1 --mode aurora-replacement --engine mysql

  # 只分析 PostgreSQL 的 RDS → Aurora 迁移
  python %(prog)s ap-southeast-1 --mode rds-to-aurora --engine postgresql

  # 优化模式（并发处理，适合大量实例）
  python %(prog)s ap-southeast-1 --optimize -w 20 -b 100

  # 指定输出文件
  python %(prog)s ap-southeast-1 -o my_report.xlsx

参数说明:
  --mode        分析模式，决定执行哪些评估:
                  rds-to-aurora       仅 RDS MySQL/PG → Aurora 迁移评估（含 RDS 机型替换列）
                  rds-replacement     仅 RDS MySQL/PG 机型替换评估（含 Aurora 迁移列）
                  aurora-replacement  仅 Aurora provisioned 实例 Graviton 换代评估
                  all                 全部执行（默认）
                注: rds-to-aurora 和 rds-replacement 共用同一次实例扫描，输出在同一份报表中。

  --engine      引擎类型: mysql / postgresql / all（默认: all）

  --optimize    启用并发模式。开启后使用 ThreadPoolExecutor 并发采集 CloudWatch 指标
                和计算成本。定价数据始终是预先批量获取的，不受此参数影响。
                适合实例数量 > 20 的场景。

  -w/--workers  并发线程数（仅在 --optimize 模式下生效）。
                每个线程处理一个实例的 CloudWatch 采集 + 成本计算。
                建议值：10-30。过高可能触发 CloudWatch API 限流。
                默认: 10，上限: 50。

  -b/--batch-size  每批处理的实例数量（仅在 --optimize 模式下生效）。
                   实例会被分成多个批次，每批内并发处理。
                   批次之间串行执行，避免同时发起过多 API 请求。
                   默认: 50。

处理流程:
  1. 批量获取整个 region 的所有 RDS/Aurora 定价（缓存到内存）
  2. 批量获取所有存储类型定价
  3. 通过 describe_db_instances 分页获取所有实例
  4. 对每个实例：采集 CloudWatch CPU/IOPS 指标 → 计算 Aurora 迁移成本 → 计算 RDS 替换成本
  5. 输出 Excel（多 Sheet）+ HTML 报告（带集群超链接和成本对比百分比）
""")
    parser.add_argument('region', help='AWS 区域名称（如 ap-southeast-1, us-east-1）')
    parser.add_argument('-o', '--output', help='输出文件名（默认: rds_aurora_analysis_<mode>_<region>_<timestamp>.xlsx）')
    parser.add_argument('--mode', choices=['rds-to-aurora', 'rds-replacement', 'aurora-replacement', 'all'], default='all',
                       help='分析模式（默认: all，全量执行）')
    parser.add_argument('--engine', choices=['mysql', 'postgresql', 'all'], default='all',
                       help='分析的引擎类型（默认: all，同时分析 MySQL 和 PostgreSQL）')
    parser.add_argument('--optimize', action='store_true',
                       help='启用并发模式，使用 ThreadPoolExecutor 并发处理实例（适合 >20 个实例）')
    parser.add_argument('-w', '--workers', type=int, default=10,
                       help='并发线程数，仅 --optimize 模式生效（默认: 10，上限: 50）')
    parser.add_argument('-b', '--batch-size', type=int, default=50,
                       help='每批处理实例数，仅 --optimize 模式生效（默认: 50）')
    
    args = parser.parse_args()
    
    # 限制最大线程数，避免 CloudWatch API 限流
    max_workers = min(args.workers, 50)
    
    try:
        analyzer = CompleteUnifiedRDSAuroraAnalyzer(
            region=args.region,
            enable_optimization=args.optimize,
            max_workers=max_workers,
            batch_size=args.batch_size,
            engine=args.engine,
            mode=args.mode
        )
        
        analyzer.run_analysis(args.output)
        
    except KeyboardInterrupt:
        print("\n分析被用户中断")
        sys.exit(1)
    except Exception as e:
        print(f"分析失败: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
