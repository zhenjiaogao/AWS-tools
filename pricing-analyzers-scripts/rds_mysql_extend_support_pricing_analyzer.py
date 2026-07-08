#!/usr/bin/env python3
"""
RDS/Aurora Extended Support 定价分析脚本
支持: RDS MySQL, RDS PostgreSQL, Aurora MySQL, Aurora PostgreSQL
功能: EOS 时间线、紧急度标注、ES enrollment 状态、按 pricing year 计费、集群汇总
"""

import boto3
import json
import csv
import argparse
from datetime import date, datetime
from typing import Dict, List, Any, Optional, Tuple
import sys

# =============================================================================
# Engine 定义 (Rule 1)
# =============================================================================

ENGINE_CONFIG = {
    'mysql': {'api_engine': 'mysql', 'pricing_engine': 'MySQL', 'family': 'RDS'},
    'postgres': {'api_engine': 'postgres', 'pricing_engine': 'PostgreSQL', 'family': 'RDS'},
    'aurora-mysql': {'api_engine': 'aurora-mysql', 'pricing_engine': 'Aurora MySQL', 'family': 'Aurora'},
    'aurora-postgresql': {'api_engine': 'aurora-postgresql', 'pricing_engine': 'Aurora PostgreSQL', 'family': 'Aurora'},
}

# EOS Fallback (Rule 2 - only used if API fails)
EOS_FALLBACK = {
    'mysql': {'5.7': ('2024-02-29', '2027-02-28'), '8.0': ('2026-07-31', '2029-07-31')},
    'postgres': {'11': ('2024-02-29', '2027-02-28'), '12': ('2025-02-28', '2028-02-28'),
                 '13': ('2026-01-31', '2029-01-31'), '14': ('2026-11-30', '2029-11-30')},
    'aurora-mysql': {'2': ('2024-10-31', '2027-10-31'), '3': ('2027-07-31', '2030-07-31')},
    'aurora-postgresql': {'13': ('2026-01-31', '2029-01-31'), '14': ('2026-11-30', '2029-11-30')},
}

# Region -> Location mapping
REGION_LOCATION = {
    'us-east-1': 'US East (N. Virginia)', 'us-east-2': 'US East (Ohio)',
    'us-west-1': 'US West (N. California)', 'us-west-2': 'US West (Oregon)',
    'eu-west-1': 'EU (Ireland)', 'eu-west-2': 'EU (London)', 'eu-west-3': 'EU (Paris)',
    'eu-central-1': 'EU (Frankfurt)', 'eu-central-2': 'Europe (Zurich)',
    'eu-north-1': 'EU (Stockholm)', 'eu-south-1': 'EU (Milan)', 'eu-south-2': 'Europe (Spain)',
    'ap-southeast-1': 'Asia Pacific (Singapore)', 'ap-southeast-2': 'Asia Pacific (Sydney)',
    'ap-southeast-3': 'Asia Pacific (Jakarta)', 'ap-southeast-4': 'Asia Pacific (Melbourne)',
    'ap-southeast-5': 'Asia Pacific (Malaysia)', 'ap-southeast-7': 'Asia Pacific (Thailand)',
    'ap-northeast-1': 'Asia Pacific (Tokyo)', 'ap-northeast-2': 'Asia Pacific (Seoul)',
    'ap-northeast-3': 'Asia Pacific (Osaka)',
    'ap-south-1': 'Asia Pacific (Mumbai)', 'ap-south-2': 'Asia Pacific (Hyderabad)',
    'ap-east-1': 'Asia Pacific (Hong Kong)',
    'ca-central-1': 'Canada (Central)', 'ca-west-1': 'Canada West (Calgary)',
    'sa-east-1': 'South America (Sao Paulo)', 'af-south-1': 'Africa (Cape Town)',
    'me-central-1': 'Middle East (UAE)', 'me-south-1': 'Middle East (Bahrain)',
    'il-central-1': 'Israel (Tel Aviv)', 'mx-central-1': 'Mexico (Central)',
}


# =============================================================================
# Helper Functions
# =============================================================================

def extract_major_version(engine: str, version: str) -> str:
    """Extract major version per Rule 2."""
    parts = version.split('.')
    if engine == 'mysql':
        return '.'.join(parts[:2]) if len(parts) >= 2 else version
    elif engine == 'postgres':
        return parts[0] if parts else version
    elif engine == 'aurora-mysql':
        # Aurora MySQL version format: 8.0.mysql_aurora.3.10.3
        # Major version is the first two segments: 8.0
        return '.'.join(parts[:2]) if len(parts) >= 2 else version
    elif engine == 'aurora-postgresql':
        # Aurora PostgreSQL version format: 14.12 or 15.4
        return parts[0] if parts else version
    return version


def get_es_pricing_year(es_start_date: date, es_end_date: date, today: date = None) -> int:
    """Rule 5: Determine ES pricing year based on ES start/end dates from API.
    Year 1: es_start → es_start + 1 year
    Year 2: es_start + 1 year → es_start + 2 years
    Year 3: es_start + 2 years → es_end
    Returns 0 if not yet in ES, 1/2/3 otherwise.
    """
    if today is None:
        today = date.today()

    if today < es_start_date:
        return 0  # Still in standard support

    if today > es_end_date:
        return 3  # Past ES end, still year 3 pricing until upgraded

    # Year boundaries using calendar years from ES start
    year2_start = date(es_start_date.year + 1, es_start_date.month, es_start_date.day)
    year3_start = date(es_start_date.year + 2, es_start_date.month, es_start_date.day)

    if today < year2_start:
        return 1
    elif today < year3_start:
        return 2
    else:
        return 3


def get_urgency(eos_date: date, today: date = None) -> str:
    """Rule 4: Urgency classification."""
    if today is None:
        today = date.today()
    if eos_date < today:
        return 'EXPIRED'
    elif eos_date.year == today.year and eos_date >= today:
        return 'URGENT'
    elif eos_date.year == today.year + 1:
        return 'APPROACHING'
    else:
        return 'SAFE'


URGENCY_SEVERITY = {'EXPIRED': 4, 'URGENT': 3, 'APPROACHING': 2, 'SAFE': 1}


# =============================================================================
# Main Analyzer Class
# =============================================================================

class ExtendedSupportAnalyzer:
    def __init__(self, region: str, profile: str = None):
        self.region = region
        session_kwargs = {'profile_name': profile} if profile else {}
        self.session = boto3.Session(**session_kwargs)
        self.rds_client = self.session.client('rds', region_name=region)
        self.pricing_client = self.session.client('pricing', region_name='us-east-1')
        self.sts_client = self.session.client('sts', region_name=region)
        self._account_id: Optional[str] = None
        self._vcpu_cache: Dict[str, int] = {}
        self._eos_cache: Dict[str, Dict[str, Dict[str, date]]] = {}  # engine -> {major_ver -> {eos, es_start, es_end}}
        self._pricing_cache: Dict[str, Dict] = {}  # engine -> {price_y12, price_y3}
        self._ri_cache: Dict[str, Dict[str, float]] = {}  # engine -> {instance_class -> hourly}

    # =========================================================================
    # Account
    # =========================================================================

    def get_account_id(self) -> str:
        if self._account_id:
            return self._account_id
        try:
            self._account_id = self.sts_client.get_caller_identity()['Account']
        except Exception as e:
            print(f"  ⚠️  获取 Account ID 失败: {e}")
            self._account_id = 'unknown'
        return self._account_id

    # =========================================================================
    # EOS Dates (Rule 2)
    # =========================================================================

    def get_eos_calendar(self, engine: str) -> Dict[str, Dict[str, date]]:
        """Get major_version -> {eos, es_start, es_end} mapping via API."""
        if engine in self._eos_cache:
            return self._eos_cache[engine]

        api_engine = ENGINE_CONFIG[engine]['api_engine']
        eos_map: Dict[str, Dict[str, date]] = {}
        used_fallback = False

        try:
            resp = self.rds_client.describe_db_major_engine_versions(Engine=api_engine)
            for ver in resp.get('DBMajorEngineVersions', []):
                major = ver.get('MajorEngineVersion', '')
                if not major:
                    continue

                entry = {'eos': None, 'es_start': None, 'es_end': None}

                for lifecycle in ver.get('SupportedEngineLifecycles', []):
                    name = lifecycle.get('LifecycleSupportName', '')
                    end_dt = lifecycle.get('LifecycleSupportEndDate')
                    start_dt = lifecycle.get('LifecycleSupportStartDate')

                    def to_date(dt):
                        if dt is None:
                            return None
                        if hasattr(dt, 'date'):
                            return dt.date()
                        if isinstance(dt, date):
                            return dt
                        if isinstance(dt, str):
                            return date.fromisoformat(dt[:10])
                        return None

                    if name == 'open-source-rds-standard-support':
                        entry['eos'] = to_date(end_dt)
                    elif name == 'open-source-rds-extended-support':
                        entry['es_start'] = to_date(start_dt)
                        entry['es_end'] = to_date(end_dt)

                if entry['eos']:
                    eos_map[major] = entry

        except Exception as e:
            print(f"  ⚠️  describe_db_major_engine_versions failed for {engine}: {e}")

        if not eos_map:
            # Use fallback
            used_fallback = True
            fallback = EOS_FALLBACK.get(engine, {})
            for major, (eos_str, es_end_str) in fallback.items():
                eos_d = date.fromisoformat(eos_str)
                es_end_d = date.fromisoformat(es_end_str)
                # ES starts the day after EOS
                es_start_d = date(eos_d.year, eos_d.month, eos_d.day)
                es_start_d = date.fromordinal(eos_d.toordinal() + 1)
                eos_map[major] = {'eos': eos_d, 'es_start': es_start_d, 'es_end': es_end_d}
            if eos_map:
                print(f"  ⚠️  WARNING: {engine} EOS dates from hardcoded fallback (API unavailable).")
                print(f"     Verify at https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/MySQL.Concepts.VersionMgmt.html")

        if not used_fallback and eos_map:
            versions_info = [f"{k}=EOS:{v['eos']}" for k, v in sorted(eos_map.items())]
            print(f"  ✅ {engine} EOS calendar: {', '.join(versions_info)}")

        self._eos_cache[engine] = eos_map
        return eos_map

    # =========================================================================
    # ES Pricing (Rule 6)
    # =========================================================================

    def get_es_pricing(self, engine: str) -> Dict[str, Optional[float]]:
        """Get Extended Support pricing for engine in current region."""
        if engine in self._pricing_cache:
            return self._pricing_cache[engine]

        pricing_engine = ENGINE_CONFIG[engine]['pricing_engine']
        location = REGION_LOCATION.get(self.region, 'US East (N. Virginia)')

        result = {'price_y12': None, 'price_y3': None}

        try:
            paginator = self.pricing_client.get_paginator('get_products')
            for page in paginator.paginate(
                ServiceCode='AmazonRDS',
                Filters=[
                    {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': pricing_engine},
                    {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
                ]
            ):
                for product_str in page['PriceList']:
                    product = json.loads(product_str)
                    attrs = product['product']['attributes']
                    usage_type = attrs.get('usagetype', '')

                    if 'ExtendedSupport' not in usage_type:
                        continue
                    if any(x in usage_type for x in [':PP', ':Gov', ':Spot']):
                        continue

                    terms = product.get('terms', {}).get('OnDemand', {})
                    for term_data in terms.values():
                        for price_data in term_data.get('priceDimensions', {}).values():
                            price_val = float(price_data.get('pricePerUnit', {}).get('USD', '0'))
                            if price_val <= 0:
                                continue
                            if 'Yr1-Yr2' in usage_type:
                                result['price_y12'] = price_val
                            elif 'Yr3' in usage_type:
                                result['price_y3'] = price_val

        except Exception as e:
            print(f"  ❌ 获取 {engine} ES 定价失败: {e}")

        # Fallback: y3 = y12 * 2
        if result['price_y12'] and not result['price_y3']:
            result['price_y3'] = result['price_y12'] * 2

        if result['price_y12']:
            print(f"  ✅ {engine} ES pricing ({self.region}): Y1-2=${result['price_y12']}, Y3=${result['price_y3']}")
        else:
            print(f"  ⚠️  {engine} ES pricing not found for {self.region}")

        self._pricing_cache[engine] = result
        return result

    # =========================================================================
    # RI Pricing (Rule 10)
    # =========================================================================

    def get_ri_pricing(self, engine: str) -> Dict[str, float]:
        """Get 1yr No Upfront RI pricing per instance class."""
        if engine in self._ri_cache:
            return self._ri_cache[engine]

        pricing_engine = ENGINE_CONFIG[engine]['pricing_engine']
        location = REGION_LOCATION.get(self.region, 'US East (N. Virginia)')
        family = ENGINE_CONFIG[engine]['family']
        ri_map: Dict[str, float] = {}

        filters = [
            {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': pricing_engine},
            {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
        ]
        if family == 'RDS':
            filters.append({'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'})

        try:
            paginator = self.pricing_client.get_paginator('get_products')
            for page in paginator.paginate(ServiceCode='AmazonRDS', Filters=filters):
                for product_str in page.get('PriceList', []):
                    product = json.loads(product_str)
                    attrs = product.get('product', {}).get('attributes', {})
                    inst_type = attrs.get('instanceType', '')
                    if not inst_type:
                        continue

                    for term in product.get('terms', {}).get('Reserved', {}).values():
                        term_attrs = term.get('termAttributes', {})
                        if (term_attrs.get('LeaseContractLength') == '1yr' and
                                term_attrs.get('PurchaseOption') == 'No Upfront'):
                            for pd in term.get('priceDimensions', {}).values():
                                if pd.get('unit') == 'Hrs':
                                    price = float(pd.get('pricePerUnit', {}).get('USD', 0))
                                    if price > 0:
                                        ri_map[inst_type] = price
                                    break

            print(f"  ✅ {engine} RI pricing: {len(ri_map)} instance types")
        except Exception as e:
            print(f"  ❌ 获取 {engine} RI 定价失败: {e}")

        self._ri_cache[engine] = ri_map
        return ri_map

    # =========================================================================
    # vCPU (Rule 9)
    # =========================================================================

    def get_vcpus(self, instance_class: str) -> int:
        if instance_class == 'db.serverless':
            return 0  # Serverless uses ACU
        if instance_class in self._vcpu_cache:
            return self._vcpu_cache[instance_class]

        ec2_type = instance_class.replace('db.', '', 1)
        try:
            ec2 = self.session.client('ec2', region_name=self.region)
            resp = ec2.describe_instance_types(InstanceTypes=[ec2_type])
            if resp['InstanceTypes']:
                vcpus = resp['InstanceTypes'][0]['VCpuInfo']['DefaultVCpus']
                self._vcpu_cache[instance_class] = vcpus
                return vcpus
        except Exception:
            pass

        # Fallback
        vcpus = self._static_vcpu(instance_class)
        self._vcpu_cache[instance_class] = vcpus
        return vcpus

    @staticmethod
    def _static_vcpu(instance_class: str) -> int:
        parts = instance_class.split('.')
        if len(parts) != 3:
            return 2
        family, size = parts[1], parts[2]
        if family.startswith('t'):
            return {'nano': 2, 'micro': 2, 'small': 2, 'medium': 2,
                    'large': 2, 'xlarge': 4, '2xlarge': 8}.get(size, 2)
        return {'small': 1, 'medium': 1, 'large': 2, 'xlarge': 4, '2xlarge': 8,
                '3xlarge': 12, '4xlarge': 16, '6xlarge': 24, '8xlarge': 32,
                '9xlarge': 36, '10xlarge': 40, '12xlarge': 48, '16xlarge': 64,
                '18xlarge': 72, '24xlarge': 96, '32xlarge': 128, '48xlarge': 192}.get(size, 2)

    # =========================================================================
    # Instance Discovery
    # =========================================================================

    def discover_instances(self, engines: List[str], instance_filter: str = None,
                           cluster_filter: str = None) -> List[Dict]:
        """Discover all relevant instances."""
        account_id = self.get_account_id()
        all_instances = []

        # Collect raw DB instances
        raw_instances = []
        if instance_filter:
            try:
                resp = self.rds_client.describe_db_instances(DBInstanceIdentifier=instance_filter)
                raw_instances = resp['DBInstances']
            except Exception as e:
                print(f"  ❌ 实例 '{instance_filter}' 查询失败: {e}")
                return []
        elif cluster_filter:
            try:
                cluster_resp = self.rds_client.describe_db_clusters(DBClusterIdentifier=cluster_filter)
                members = cluster_resp['DBClusters'][0].get('DBClusterMembers', [])
                for m in members:
                    resp = self.rds_client.describe_db_instances(DBInstanceIdentifier=m['DBInstanceIdentifier'])
                    raw_instances.extend(resp['DBInstances'])
            except Exception:
                # Try as primary instance with replicas
                paginator = self.rds_client.get_paginator('describe_db_instances')
                for page in paginator.paginate():
                    for inst in page['DBInstances']:
                        if (inst['DBInstanceIdentifier'] == cluster_filter or
                                inst.get('ReadReplicaSourceDBInstanceIdentifier') == cluster_filter):
                            raw_instances.append(inst)
        else:
            paginator = self.rds_client.get_paginator('describe_db_instances')
            for page in paginator.paginate():
                raw_instances.extend(page['DBInstances'])

        # Filter by engine
        api_engines = [ENGINE_CONFIG[e]['api_engine'] for e in engines]

        for db_inst in raw_instances:
            inst_engine = db_inst['Engine'].lower()
            if inst_engine not in api_engines:
                continue

            # Map back to our engine key
            engine_key = next(k for k, v in ENGINE_CONFIG.items() if v['api_engine'] == inst_engine)

            # ES enrollment (Rule 3)
            es_field = db_inst.get('EngineLifecycleSupport')
            if not es_field:
                # For Aurora, check cluster level
                cluster_id = db_inst.get('DBClusterIdentifier')
                if cluster_id:
                    try:
                        cr = self.rds_client.describe_db_clusters(DBClusterIdentifier=cluster_id)
                        es_field = cr['DBClusters'][0].get('EngineLifecycleSupport')
                    except Exception:
                        pass

            if es_field == 'open-source-rds-extended-support':
                es_label = 'enabled'
            elif es_field == 'open-source-rds-extended-support-disabled':
                es_label = 'disabled'
            else:
                es_label = 'default'

            # Architecture & multiplier (Rule 7)
            cluster_id = db_inst.get('DBClusterIdentifier', '')
            multi_az = db_inst.get('MultiAZ', False)

            if multi_az and not cluster_id:
                # Classic Multi-AZ Instance (standby hidden)
                multiplier = 2
                architecture = 'multi_az_instance'
            elif cluster_id:
                # Member of a cluster — determine type
                multiplier = 1
                if ENGINE_CONFIG[engine_key]['family'] == 'Aurora':
                    architecture = 'aurora_cluster'
                else:
                    # RDS MySQL/PG cluster — check if it's Multi-AZ DB Cluster (3 nodes, 3 AZs)
                    try:
                        cr = self.rds_client.describe_db_clusters(DBClusterIdentifier=cluster_id)
                        cluster_info = cr['DBClusters'][0]
                        member_count = len(cluster_info.get('DBClusterMembers', []))
                        cluster_multi_az = cluster_info.get('MultiAZ', False)
                        if cluster_multi_az and member_count == 3:
                            architecture = 'multi_az_db_cluster'
                        else:
                            architecture = 'rds_cluster'
                    except Exception:
                        architecture = 'rds_cluster'
            elif db_inst.get('ReadReplicaSourceDBInstanceIdentifier') or db_inst.get('ReadReplicaSourceDBClusterIdentifier'):
                multiplier = 1
                architecture = 'read_replica'
            else:
                multiplier = 1
                architecture = 'single_az'

            # Source
            source = db_inst.get('ReadReplicaSourceDBInstanceIdentifier', '') or \
                     db_inst.get('ReadReplicaSourceDBClusterIdentifier', '')

            # Cluster group (Rule 12)
            if cluster_id:
                group_key = cluster_id
            elif source:
                group_key = source
            else:
                group_key = db_inst['DBInstanceIdentifier']

            # Version info
            version = db_inst['EngineVersion']
            major = extract_major_version(engine_key, version)

            # Minor version status
            minor_status = 'deprecated' if db_inst.get('Status') == 'deprecated' else 'current'

            all_instances.append({
                'account_id': account_id,
                'region': self.region,
                'engine': engine_key,
                'cluster_group': group_key,
                'db_instance_identifier': db_inst['DBInstanceIdentifier'],
                'db_cluster_identifier': cluster_id,
                'source': source,
                'engine_version': version,
                'major_version': major,
                'minor_version_status': minor_status,
                'es_enrollment': es_label,
                'architecture': architecture,
                'instance_class': db_inst['DBInstanceClass'],
                'status': db_inst.get('DBInstanceStatus', ''),
                'vcpus': self.get_vcpus(db_inst['DBInstanceClass']),
                'multiplier': multiplier,
                'global_cluster_id': '',  # Will be populated later in global mode
            })

        print(f"  ✅ 找到 {len(all_instances)} 个实例")
        return all_instances

    # =========================================================================
    # Analyze & Export
    # =========================================================================

    def analyze(self, engines: List[str], instance_filter=None, cluster_filter=None,
                output_file=None, global_mode=False):
        engine_list = ', '.join(engines)
        mode = '全部实例'
        if instance_filter:
            mode = f'实例: {instance_filter}'
        elif cluster_filter:
            mode = f'集群: {cluster_filter}'
        elif global_mode:
            mode = '全部实例 + Global Database 跨 Region 汇总'

        print(f"\n{'='*65}")
        print(f"  RDS/Aurora Extended Support 费用分析")
        print(f"  Account: {self.get_account_id()}")
        print(f"  Region:  {self.region}")
        print(f"  Engines: {engine_list}")
        print(f"  范围:    {mode}")
        print(f"{'='*65}\n")

        # Fetch pricing & EOS for each engine
        for eng in engines:
            self.get_eos_calendar(eng)
            self.get_es_pricing(eng)
            self.get_ri_pricing(eng)

        # Discover instances in current region
        instances = self.discover_instances(engines, instance_filter, cluster_filter)

        # Global mode: discover global databases and scan secondary regions
        global_db_map = {}  # global_cluster_id -> {members: [{region, cluster_id}]}
        extra_region_instances = []

        if global_mode and not instance_filter and not cluster_filter:
            global_db_map, extra_region_instances = self._discover_global_databases(engines)
            if extra_region_instances:
                instances.extend(extra_region_instances)

        # Tag instances with global_cluster_identifier
        # Build cluster_id -> global_db mapping
        cluster_to_global = {}
        for gdb_id, gdb_info in global_db_map.items():
            for member in gdb_info['members']:
                cluster_to_global[member['cluster_id']] = gdb_id

        for inst in instances:
            cid = inst.get('db_cluster_identifier', '') or inst.get('cluster_group', '')
            inst['global_cluster_id'] = cluster_to_global.get(cid, '')

        if not instances:
            print("\n未找到匹配的实例。")
            return

        # Calculate fees
        today = date.today()
        results = []

        for inst in instances:
            eng = inst['engine']
            eos_map = self._eos_cache.get(eng, {})

            # Use per-instance pricing if from secondary region, otherwise use cached
            if '_pricing' in inst:
                pricing = inst.pop('_pricing')
            else:
                pricing = self._pricing_cache.get(eng, {'price_y12': None, 'price_y3': None})

            if '_ri_map' in inst:
                ri_map = inst.pop('_ri_map')
            else:
                ri_map = self._ri_cache.get(eng, {})

            # EOS date for this major version
            version_info = eos_map.get(inst['major_version'])
            if version_info:
                eos_date = version_info['eos']
                es_start = version_info['es_start']
                es_end = version_info['es_end']
                eos_str = eos_date.isoformat() if eos_date else 'N/A'
                urgency = get_urgency(eos_date, today) if eos_date else 'UNKNOWN'
                if es_start and es_end:
                    es_year = get_es_pricing_year(es_start, es_end, today)
                else:
                    es_year = 0
            else:
                eos_str = 'N/A'
                urgency = 'UNKNOWN'
                es_year = 0

            # Pricing year selection (Rule 8)
            price_y12 = pricing['price_y12']
            price_y3 = pricing['price_y3']

            vcpus = inst['vcpus']
            mult = inst['multiplier']
            is_serverless = (inst['instance_class'] == 'db.serverless')

            if is_serverless:
                # Serverless: per-ACU pricing, current vCPU logic not applicable
                y1_mrr = 'N/A (Serverless)'
                y3_mrr = 'N/A (Serverless)'
                ri_mrr = 'N/A (Serverless)'
                es_vs_ri_y1 = 'N/A'
                es_vs_ri_y3 = 'N/A'
            else:
                # Projected fees (always calculate all years)
                y1_mrr = round((price_y12 or 0) * vcpus * mult * 730, 2)
                y3_mrr = round((price_y3 or 0) * vcpus * mult * 730, 2)

                # RI baseline
                ri_hourly = ri_map.get(inst['instance_class'], 0) * mult
                ri_mrr = round(ri_hourly * 730, 2)

                # Percentage for each year
                es_vs_ri_y1 = round((y1_mrr / ri_mrr) * 100, 1) if ri_mrr > 0 else None
                es_vs_ri_y3 = round((y3_mrr / ri_mrr) * 100, 1) if ri_mrr > 0 else None

            results.append({
                **inst,
                'major_version_eos': eos_str,
                'urgency': urgency,
                'es_pricing_year': es_year if es_year > 0 else 'N/A',
                'ri_1yr_noup_mrr_usd': ri_mrr,
                'es_year1_2_mrr_usd': y1_mrr,
                'es_year1_2_vs_ri_pct(%)': es_vs_ri_y1,
                'es_year3_mrr_usd': y3_mrr,
                'es_year3_vs_ri_pct(%)': es_vs_ri_y3,
            })

        # Export
        if not output_file:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"rds_extended_support_report_{self.region}_{ts}.xlsx"

        self._export(results, output_file, global_mode)
        self._print_summary(results, global_mode)

    def _discover_global_databases(self, engines: List[str]) -> Tuple[Dict, List[Dict]]:
        """Discover Global Databases and scan secondary region instances."""
        global_db_map = {}
        extra_instances = []

        print(f"\n  🌐 扫描 Global Databases...")
        try:
            resp = self.rds_client.describe_global_clusters()
            global_clusters = resp.get('GlobalClusters', [])

            if not global_clusters:
                print(f"  ℹ️  未发现 Global Database")
                return {}, []

            print(f"  ✅ 发现 {len(global_clusters)} 个 Global Database")

            # Collect secondary regions that differ from primary
            secondary_regions = set()
            for gc in global_clusters:
                gdb_id = gc['GlobalClusterIdentifier']
                members = []
                for member in gc.get('GlobalClusterMembers', []):
                    cluster_arn = member.get('DBClusterIdentifier', '')
                    is_writer = member.get('IsWriter', False)
                    # Extract region from ARN: arn:aws:rds:<region>:<account>:cluster:<name>
                    arn_parts = cluster_arn.split(':')
                    member_region = arn_parts[3] if len(arn_parts) > 3 else self.region
                    cluster_name = arn_parts[-1] if arn_parts else ''
                    members.append({
                        'region': member_region,
                        'cluster_id': cluster_name,
                        'cluster_arn': cluster_arn,
                        'is_writer': is_writer,
                    })
                    if member_region != self.region:
                        secondary_regions.add(member_region)

                global_db_map[gdb_id] = {'members': members}
                print(f"    {gdb_id}: {len(members)} members across {len(set(m['region'] for m in members))} regions")

            # Scan secondary regions
            for sec_region in secondary_regions:
                print(f"\n  🌐 扫描 secondary region: {sec_region}")
                sec_analyzer = ExtendedSupportAnalyzer(sec_region, profile=self.session.profile_name)

                # Fetch pricing for secondary region
                for eng in engines:
                    sec_analyzer.get_eos_calendar(eng)
                    sec_analyzer.get_es_pricing(eng)
                    sec_analyzer.get_ri_pricing(eng)

                # Get instances in secondary region
                sec_instances = sec_analyzer.discover_instances(engines)

                # Copy caches so fee calculation works
                for eng in engines:
                    cache_key = f"{sec_region}_{eng}"
                    if eng in sec_analyzer._eos_cache:
                        self._eos_cache.setdefault(eng, {}).update(sec_analyzer._eos_cache[eng])
                    # Store secondary region pricing separately on each instance
                    for inst in sec_instances:
                        if inst['engine'] == eng:
                            inst['_pricing'] = sec_analyzer._pricing_cache.get(eng, {'price_y12': None, 'price_y3': None})
                            inst['_ri_map'] = sec_analyzer._ri_cache.get(eng, {})

                extra_instances.extend(sec_instances)

        except Exception as e:
            print(f"  ⚠️  Global Database 扫描失败: {e}")

        return global_db_map, extra_instances

    def _export(self, results: List[Dict], output_file: str, global_mode: bool = False):
        # Change extension to .xlsx
        if output_file.endswith('.csv'):
            output_file = output_file.replace('.csv', '.xlsx')
        elif not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Header style
        header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='232F3E', end_color='232F3E', fill_type='solid')

        def write_sheet(ws, headers, rows):
            # Write headers
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            # Write data
            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    val = row.get(h, '')
                    if val is None:
                        val = 'N/A'
                    cell = ws.cell(row=ri, column=ci, value=val)
                    # Apply number format for money columns
                    if isinstance(val, (int, float)) and 'mrr' in h.lower():
                        cell.number_format = '$#,##0.00'
                    elif isinstance(val, (int, float)) and 'pct' in h.lower():
                        cell.number_format = '#,##0.0"%"'
            # Auto-width based on content
            for ci, h in enumerate(headers, 1):
                max_len = len(str(h))
                for ri in range(2, len(rows) + 2):
                    cell_val = ws.cell(row=ri, column=ci).value
                    if cell_val is not None:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 35)
            # Freeze header
            ws.freeze_panes = 'A2'

        # Sheet 1: Instance Detail
        ws_detail = wb.active
        ws_detail.title = 'Instance Detail'
        detail_fields = [
            'account_id', 'region', 'engine', 'global_cluster_id', 'cluster_group',
            'db_instance_identifier', 'source',
            'engine_version', 'major_version', 'major_version_eos',
            'architecture', 'instance_class', 'status', 'vcpus',
            'ri_1yr_noup_mrr_usd',
            'es_enrollment', 'urgency', 'es_pricing_year',
            'es_year1_2_mrr_usd', 'es_year1_2_vs_ri_pct(%)',
            'es_year3_mrr_usd', 'es_year3_vs_ri_pct(%)',
        ]
        detail_rows = []
        for r in results:
            row = {k: (r.get(k) if r.get(k) is not None else 'N/A') for k in detail_fields}
            detail_rows.append(row)
        write_sheet(ws_detail, detail_fields, detail_rows)

        # Apply conditional coloring based on es_pricing_year
        green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        red_fill = PatternFill(start_color='FCE4E4', end_color='FCE4E4', fill_type='solid')

        # Urgency coloring
        urgency_expired_fill = PatternFill(start_color='CC3333', end_color='CC3333', fill_type='solid')
        urgency_urgent_fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
        urgency_expired_font = Font(bold=True, color='FFFFFF')
        urgency_urgent_font = Font(bold=True, color='FFFFFF')

        # Find column indices
        y12_mrr_col = detail_fields.index('es_year1_2_mrr_usd') + 1
        y12_pct_col = detail_fields.index('es_year1_2_vs_ri_pct(%)') + 1
        y3_mrr_col = detail_fields.index('es_year3_mrr_usd') + 1
        y3_pct_col = detail_fields.index('es_year3_vs_ri_pct(%)') + 1
        urgency_col = detail_fields.index('urgency') + 1

        for ri, r in enumerate(results, 2):  # start from row 2
            es_yr = r.get('es_pricing_year', 'N/A')
            if es_yr in (1, 2):
                # Highlight Y1-2 columns green
                for col in [y12_mrr_col, y12_pct_col]:
                    ws_detail.cell(row=ri, column=col).fill = green_fill
            elif es_yr == 3:
                # Highlight Y3 columns red
                for col in [y3_mrr_col, y3_pct_col]:
                    ws_detail.cell(row=ri, column=col).fill = red_fill

            # Urgency cell coloring
            urgency_val = r.get('urgency', '')
            urgency_cell = ws_detail.cell(row=ri, column=urgency_col)
            if urgency_val == 'EXPIRED':
                urgency_cell.fill = urgency_expired_fill
                urgency_cell.font = urgency_expired_font
            elif urgency_val == 'URGENT':
                urgency_cell.fill = urgency_urgent_fill
                urgency_cell.font = urgency_urgent_font

        wb.save(output_file)
        print(f"\n  📄 报告已保存: {output_file}")

    def _cluster_summary(self, results: List[Dict]) -> List[Dict]:
        groups: Dict[str, List[Dict]] = {}
        for r in results:
            # Group by region + engine + cluster_group to avoid cross-engine collisions
            key = f"{r['region']}::{r['engine']}::{r['cluster_group']}"
            groups.setdefault(key, []).append(r)

        def safe_sum(members, field):
            total = 0
            for m in members:
                val = m.get(field, 0)
                if isinstance(val, (int, float)):
                    total += val
            return round(total, 2)

        summary = []
        for key, members in groups.items():
            worst_urgency = max(members, key=lambda m: URGENCY_SEVERITY.get(m['urgency'], 0))['urgency']
            summary.append({
                'cluster_group': members[0]['cluster_group'],
                'region': members[0]['region'],
                'engine': members[0]['engine'],
                'global_cluster_id': members[0].get('global_cluster_id', ''),
                'total_instances': len(members),
                'total_vcpus': sum(m['vcpus'] * m['multiplier'] for m in members),
                'architecture_desc': ' + '.join(sorted(set(m['architecture'] for m in members))),
                'urgency': worst_urgency,
                'ri_total_mrr_usd': safe_sum(members, 'ri_1yr_noup_mrr_usd'),
                'es_year1_2_total_mrr_usd': safe_sum(members, 'es_year1_2_mrr_usd'),
                'es_year3_total_mrr_usd': safe_sum(members, 'es_year3_mrr_usd'),
            })
        return summary

    def _global_db_summary(self, results: List[Dict]) -> List[Dict]:
        """Aggregate by global_cluster_id for cross-region Global DB summary."""
        global_groups: Dict[str, List[Dict]] = {}
        for r in results:
            gid = r.get('global_cluster_id', '')
            if gid:
                global_groups.setdefault(gid, []).append(r)

        if not global_groups:
            return []

        def safe_sum(members, field):
            total = 0
            for m in members:
                val = m.get(field, 0)
                if isinstance(val, (int, float)):
                    total += val
            return round(total, 2)

        summary = []
        for gid, members in global_groups.items():
            worst_urgency = max(members, key=lambda m: URGENCY_SEVERITY.get(m['urgency'], 0))['urgency']
            regions = set(m['region'] for m in members)
            clusters = set(m['cluster_group'] for m in members)
            summary.append({
                'global_cluster_id': gid,
                'total_regions': len(regions),
                'total_clusters': len(clusters),
                'total_instances': len(members),
                'total_vcpus': sum(m['vcpus'] * m['multiplier'] for m in members),
                'urgency': worst_urgency,
                'es_year1_2_total_mrr_usd': safe_sum(members, 'es_year1_2_mrr_usd'),
                'es_year3_total_mrr_usd': safe_sum(members, 'es_year3_mrr_usd'),
            })
        return summary

    def _print_summary(self, results: List[Dict], global_mode: bool = False):
        summary = self._cluster_summary(results)
        print(f"\n{'='*90}")
        print(f"  集群级 Extended Support 费用汇总")
        print(f"{'='*90}")
        print(f"{'Cluster':<25} {'Region':<16} {'Eng':<8} {'Urg':<11} {'Nodes':>5} {'vCPUs':>6} "
              f"{'Y1-2 MRR':>10} {'Y3 MRR':>10}")
        print(f"{'-'*90}")

        total_y12 = total_y3 = 0
        for s in summary:
            y12 = s['es_year1_2_total_mrr_usd']
            y3 = s['es_year3_total_mrr_usd']
            total_y12 += y12 if isinstance(y12, (int, float)) else 0
            total_y3 += y3 if isinstance(y3, (int, float)) else 0

            y12_str = f"${y12:>8.0f}" if isinstance(y12, (int, float)) else f"{'N/A':>9}"
            y3_str = f"${y3:>8.0f}" if isinstance(y3, (int, float)) else f"{'N/A':>9}"

            print(f"{s['cluster_group']:<25} {s['region']:<16} {s['engine']:<8} {s['urgency']:<11} "
                  f"{s['total_instances']:>5} {s['total_vcpus']:>6} "
                  f"{y12_str} {y3_str}")

        print(f"{'-'*90}")
        print(f"{'TOTAL':<25} {'':16} {'':8} {'':11} {len(results):>5} {'':>6} "
              f"${total_y12:>8.0f} ${total_y3:>8.0f}")
        print(f"{'='*90}")

        # Global DB summary
        if global_mode:
            global_summary = self._global_db_summary(results)
            if global_summary:
                print(f"\n{'='*90}")
                print(f"  Global Database 跨 Region 汇总")
                print(f"{'='*90}")
                print(f"{'Global DB':<35} {'Regions':>7} {'Clusters':>8} {'Nodes':>6} "
                      f"{'Y1-2 MRR':>10} {'Y3 MRR':>10}")
                print(f"{'-'*90}")
                for gs in global_summary:
                    print(f"{gs['global_cluster_id']:<35} {gs['total_regions']:>7} "
                          f"{gs['total_clusters']:>8} {gs['total_instances']:>6} "
                          f"${gs['es_year1_2_total_mrr_usd']:>8.0f} "
                          f"${gs['es_year3_total_mrr_usd']:>8.0f}")
                print(f"{'='*90}\n")


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='RDS/Aurora Extended Support 费用分析工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python rds_extended_support_pricing_analyzer.py us-east-1 --profile my-profile
  python rds_extended_support_pricing_analyzer.py us-east-1 --engine mysql
  python rds_extended_support_pricing_analyzer.py us-east-1 --instance database-1
  python rds_extended_support_pricing_analyzer.py us-east-1 --cluster my-aurora-cluster
        """
    )
    parser.add_argument('region', help='AWS region (e.g., us-east-1)')
    parser.add_argument('--profile', help='AWS CLI profile name')
    parser.add_argument('--engine', default='all',
                        choices=['mysql', 'postgres', 'aurora-mysql', 'aurora-postgresql', 'all'],
                        help='Engine filter (default: all)')
    parser.add_argument('--instance', help='Specific DB instance identifier')
    parser.add_argument('--cluster', help='Specific DB cluster identifier')
    parser.add_argument('--global', dest='global_mode', action='store_true',
                        help='Scan Global Databases across all member regions')
    parser.add_argument('-o', '--output', help='Output CSV filename')

    args = parser.parse_args()

    engines = list(ENGINE_CONFIG.keys()) if args.engine == 'all' else [args.engine]

    try:
        analyzer = ExtendedSupportAnalyzer(args.region, profile=args.profile)
        analyzer.analyze(engines, instance_filter=args.instance,
                         cluster_filter=args.cluster, output_file=args.output,
                         global_mode=args.global_mode)
    except KeyboardInterrupt:
        print("\n用户取消。")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 执行失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
